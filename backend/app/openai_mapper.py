from __future__ import annotations

import hashlib
import json
import os
from pathlib import Path
from typing import Any, Dict, List, Sequence

from openpyxl import load_workbook


STANDARD_FIELDS = [
    "section_name",
    "work_description",
    "quantity",
    "unit",
    "unit_price",
    "planned_cost",
    "actual_completed_cost",
    "remaining_cost",
    "planned_execution",
    "actual_execution",
    "planned_start",
    "planned_finish",
    "actual_start",
    "actual_finish",
    "duration_days",
    "workforce_current",
    "workforce_required",
    "material_name",
    "ordered_quantity",
    "delivered_quantity",
    "not_mapped",
]

SHEET_TYPES = [
    "cost_estimate",
    "progress_payment",
    "schedule",
    "workforce",
    "procurement",
    "report",
    "supporting_document",
    "unknown",
]


def is_mapping_enabled() -> bool:
    return os.getenv("OPENAI_MAPPING_ENABLED", "false").strip().lower() in {"1", "true", "yes", "on"}


def mapping_model_name() -> str:
    return os.getenv("OPENAI_MAPPING_MODEL", "gpt-4.1-mini").strip() or "gpt-4.1-mini"


def should_call_openai(confidence: int) -> bool:
    try:
        threshold = int(os.getenv("OPENAI_MAPPING_CONFIDENCE_THRESHOLD", "85"))
    except Exception:
        threshold = 85
    return is_mapping_enabled() and bool(os.getenv("OPENAI_API_KEY")) and confidence < threshold


def build_workbook_context(paths: Sequence[Path], sheet_profiles: Sequence[Dict[str, Any]], analysis_type: str) -> Dict[str, Any]:
    """Build a small, privacy-aware mapping context.

    Only sheet names, likely headers, and a few sample rows are extracted. Full
    workbooks are never sent to the model.
    """
    profile_lookup = {}
    for profile in sheet_profiles:
        profile_lookup[(profile.get("file_name"), profile.get("sheet_name"))] = profile

    sheets: List[Dict[str, Any]] = []
    for path in paths:
        if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            continue
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
        except Exception:
            continue
        for ws in wb.worksheets:
            rows: List[List[Any]] = []
            for row in ws.iter_rows(max_row=40, values_only=True):
                clean = [_clean_cell(v) for v in row]
                if any(v not in (None, "") for v in clean):
                    rows.append(clean)
            if not rows:
                continue
            profile = profile_lookup.get((path.name, ws.title), {})
            header_idx = _choose_header_row(rows, profile)
            headers = _trim_row(rows[header_idx] if header_idx is not None else rows[0])
            sample_rows = []
            start = (header_idx + 1) if header_idx is not None else 1
            for row in rows[start:start + 5]:
                trimmed = _trim_row(row)
                if trimmed:
                    sample_rows.append(trimmed)
            largest_numbers = _largest_numbers(rows)
            sheets.append({
                "file_name": path.name,
                "sheet_name": ws.title,
                "detected_type": profile.get("detected_type"),
                "parser_confidence": profile.get("confidence"),
                "mapped_columns": profile.get("mapped_columns", {}),
                "headers": headers,
                "sample_rows": sample_rows,
                "largest_numbers": largest_numbers[:6],
            })
    return {
        "analysis_type": analysis_type,
        "standard_fields": STANDARD_FIELDS,
        "sheet_types": SHEET_TYPES,
        "sheets": sheets[:12],
    }


def run_assisted_mapping(context: Dict[str, Any]) -> Dict[str, Any]:
    """Call OpenAI for assisted sheet/column mapping.

    The function fails closed: if the API is disabled, unavailable or returns an
    invalid payload, the backend still works using the deterministic parser.
    """
    if not is_mapping_enabled():
        return {"enabled": False, "reason": "OPENAI_MAPPING_ENABLED is not true"}
    if not os.getenv("OPENAI_API_KEY"):
        return {"enabled": False, "reason": "OPENAI_API_KEY is not configured"}
    if not context.get("sheets"):
        return {"enabled": False, "reason": "No sheet context was available for mapping"}

    cache_key = _context_hash(context)
    cached = _load_cached_mapping(cache_key)
    if cached:
        cached["cached"] = True
        return cached

    try:
        from openai import OpenAI  # type: ignore
    except Exception as exc:
        return {"enabled": False, "reason": f"OpenAI SDK not installed: {exc}"}

    client = OpenAI()
    system = (
        "You are a construction project-control column mapping assistant. "
        "Map sheet names, headers and a few sample rows into standard construction fields. "
        "Do not calculate final totals, execution percentages, risk scores or commercial results. "
        "Return only JSON. Use not_mapped when uncertain."
    )
    user_payload = {
        "task": "Map construction sheets and columns for DevBareun preflight.",
        "rules": [
            "Never infer final monetary totals.",
            "Never calculate actual execution percentage.",
            "Only classify sheet type and likely column meaning.",
            "Use confidence values from 0 to 1.",
            "Ask user questions for missing or unclear fields.",
        ],
        "context": context,
    }
    try:
        response = client.chat.completions.create(
            model=mapping_model_name(),
            temperature=0,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
            ],
        )
        content = response.choices[0].message.content or "{}"
        payload = json.loads(content)
        normalized = _normalize_mapping_payload(payload)
        normalized["enabled"] = True
        normalized["model"] = mapping_model_name()
        normalized["cached"] = False
        normalized["context_hash"] = cache_key
        _save_cached_mapping(cache_key, normalized)
        return normalized
    except Exception:
        return {
            "enabled": True,
            "error": "Assisted mapping is temporarily unavailable.",
            "model": mapping_model_name(),
            "context_hash": cache_key,
        }


def _normalize_mapping_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    sheets = payload.get("sheets") or payload.get("sheet_mappings") or []
    normalized_sheets: List[Dict[str, Any]] = []
    for item in sheets if isinstance(sheets, list) else []:
        if not isinstance(item, dict):
            continue
        sheet_type = item.get("sheet_type") or item.get("type") or "unknown"
        if sheet_type not in SHEET_TYPES:
            sheet_type = "unknown"
        columns = item.get("columns") if isinstance(item.get("columns"), dict) else {}
        filtered_columns = {}
        for column, field in columns.items():
            filtered_columns[str(column)] = field if field in STANDARD_FIELDS else "not_mapped"
        normalized_sheets.append({
            "file_name": item.get("file_name"),
            "sheet_name": item.get("sheet_name"),
            "sheet_type": sheet_type,
            "confidence": _safe_confidence(item.get("confidence")),
            "columns": filtered_columns,
            "reason": str(item.get("reason") or "")[:400],
        })
    questions = payload.get("user_questions") or payload.get("questions") or []
    if not isinstance(questions, list):
        questions = []
    missing = payload.get("missing_fields") or []
    if not isinstance(missing, list):
        missing = []
    warnings = payload.get("warnings") or []
    if not isinstance(warnings, list):
        warnings = []
    return {
        "sheet_mappings": normalized_sheets,
        "missing_fields": [str(v) for v in missing[:12]],
        "warnings": [str(v)[:300] for v in warnings[:12]],
        "user_questions": [q if isinstance(q, dict) else {"question": str(q)} for q in questions[:12]],
        "confidence": _safe_confidence(payload.get("confidence")),
    }


def _choose_header_row(rows: Sequence[Sequence[Any]], profile: Dict[str, Any]) -> int | None:
    header_row = profile.get("header_row")
    if isinstance(header_row, int) and header_row > 0 and header_row - 1 < len(rows):
        return header_row - 1
    best_idx = None
    best_score = -1
    for idx, row in enumerate(rows[:25]):
        text_cells = sum(1 for v in row if isinstance(v, str) and v.strip())
        num_cells = sum(1 for v in row if isinstance(v, (int, float)) and not isinstance(v, bool))
        filled = sum(1 for v in row if v not in (None, ""))
        score = text_cells * 3 + filled - num_cells
        if score > best_score and filled >= 2:
            best_idx = idx
            best_score = score
    return best_idx


def _trim_row(row: Sequence[Any]) -> List[Any]:
    result: List[Any] = []
    for value in row[:18]:
        clean = _clean_cell(value)
        if clean not in (None, ""):
            result.append(clean)
        else:
            result.append("")
    while result and result[-1] == "":
        result.pop()
    return result


def _largest_numbers(rows: Sequence[Sequence[Any]]) -> List[float]:
    nums: List[float] = []
    for row in rows[:60]:
        for value in row[:20]:
            if isinstance(value, (int, float)) and not isinstance(value, bool) and 0 < float(value) < 1_000_000_000:
                nums.append(float(value))
    return sorted(set(round(v, 2) for v in nums), reverse=True)[:10]


def _clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return " ".join(value.split()).strip()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return round(float(value), 4)
    return str(value)


def _safe_confidence(value: Any) -> float:
    try:
        number = float(value)
        if number > 1:
            number = number / 100
        return max(0.0, min(1.0, round(number, 2)))
    except Exception:
        return 0.0


def _context_hash(context: Dict[str, Any]) -> str:
    raw = json.dumps(context, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:24]


def _cache_dir() -> Path:
    base = Path(__file__).resolve().parent.parent / "data" / "mapping_cache"
    base.mkdir(parents=True, exist_ok=True)
    return base


def _load_cached_mapping(cache_key: str) -> Dict[str, Any] | None:
    path = _cache_dir() / f"{cache_key}.json"
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _save_cached_mapping(cache_key: str, payload: Dict[str, Any]) -> None:
    try:
        (_cache_dir() / f"{cache_key}.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass
