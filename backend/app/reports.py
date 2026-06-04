from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List
from xml.sax.saxutils import escape as xml_escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A3, A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


_FONT_READY = False
_REGULAR_FONT = "Helvetica"
_BOLD_FONT = "Helvetica-Bold"


def _register_unicode_fonts() -> None:
    """Register a Unicode-capable font when available.

    ReportLab's built-in Helvetica cannot render Azerbaijani/Turkish letters such as
    Ə, ə, İ, ı, Ş, ş, Ğ, ğ. We do not ship font files in the project; instead we
    look for common system fonts on Linux/Railway and fall back safely if none exist.
    """
    global _FONT_READY, _REGULAR_FONT, _BOLD_FONT
    if _FONT_READY:
        return

    regular_candidates = [
        os.getenv("DEVBAREUN_PDF_FONT"),
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
        "/usr/local/share/fonts/DejaVuSans.ttf",
    ]
    bold_candidates = [
        os.getenv("DEVBAREUN_PDF_FONT_BOLD"),
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf",
        "/usr/local/share/fonts/DejaVuSans-Bold.ttf",
    ]

    # Railway's slim Python image may not include OS fonts. If matplotlib is
    # installed, it carries DejaVu Sans inside the wheel, which supports the
    # Azerbaijani characters used in smeta/F-2 files, including Ə/ə.
    try:
        import matplotlib  # type: ignore

        mpl_font_dir = Path(matplotlib.__file__).resolve().parent / "mpl-data" / "fonts" / "ttf"
        regular_candidates.append(str(mpl_font_dir / "DejaVuSans.ttf"))
        bold_candidates.append(str(mpl_font_dir / "DejaVuSans-Bold.ttf"))
    except Exception:
        pass

    # ReportLab bundles Bitstream Vera; it is useful fallback but does not cover
    # every Azerbaijani glyph, so it is intentionally checked after DejaVu/Noto.
    try:
        import reportlab  # type: ignore

        rl_font_dir = Path(reportlab.__file__).resolve().parent / "fonts"
        regular_candidates.append(str(rl_font_dir / "Vera.ttf"))
        bold_candidates.append(str(rl_font_dir / "VeraBd.ttf"))
    except Exception:
        pass

    regular_path = next((p for p in regular_candidates if p and Path(p).exists()), None)
    bold_path = next((p for p in bold_candidates if p and Path(p).exists()), None)

    if regular_path:
        pdfmetrics.registerFont(TTFont("DevBareunUnicode", regular_path))
        _REGULAR_FONT = "DevBareunUnicode"
    if bold_path:
        pdfmetrics.registerFont(TTFont("DevBareunUnicodeBold", bold_path))
        _BOLD_FONT = "DevBareunUnicodeBold"
    elif regular_path:
        _BOLD_FONT = _REGULAR_FONT

    _FONT_READY = True


def _text(value: Any, fallback: str = "Not available") -> str:
    if value is None or value == "":
        return fallback
    return str(value)


def _paragraph(value: Any, style: ParagraphStyle) -> Paragraph:
    return Paragraph(xml_escape(_text(value)), style)


def _paper_size(paper: str | None):
    value = str(paper or "a4").lower().strip()
    if value in {"a3", "a3-landscape", "a3_landscape"}:
        return landscape(A3)
    return A4




_LABELS = {
    "en": {
        "report_title": "DevBareun Project Report",
        "project": "Project", "report_id": "Report ID", "status": "Status", "currency": "Currency",
        "confidence": "Dashboard confidence", "dashboard_type": "Dashboard type", "dashboard_view": "Dashboard View",
        "selected_dashboard": "Selected analysis dashboard",
        "report_follows": "This report follows the selected analysis dashboard.",
        "metric": "Metric", "value": "Value", "note": "Note",
        "key_indicators": "Key Indicators", "planned_execution": "Planned execution", "actual_execution": "Actual execution",
        "progress_gap": "Progress gap", "delay": "Delay", "total_cost": "Total cost / cost estimate",
        "planned_cost": "Planned cost", "actual_cost": "Actual cost", "cost_variance": "Cost variance",
        "workforce": "Workforce", "risk_score": "Risk score", "risk_level": "Risk level",
        "baseline_finish": "Baseline finish", "estimated_finish": "Estimated finish",
        "executive_summary": "Executive Summary", "recommended_actions": "Recommended Actions",
        "no_action": "No recommended action was generated from the available data.",
        "risk_register": "Risk Register", "risk": "Risk", "level": "Level", "reason": "Reason", "action": "Action",
        "data_quality": "Data Quality Notes", "sheet_profiles": "Detected Sheet Profiles",
        "file": "File", "sheet": "Sheet", "type": "Type", "rows": "Rows", "not_available": "Not available",
        "forecast": "Forecast", "unit": "Unit", "panel": "Panel", "kpi": "KPI", "warning": "Warning",
        "header_row": "Header Row", "mapped_columns": "Mapped Columns", "signals": "Signals", "days": "days", "required": "required"
    },
    "az": {
        "report_title": "DevBareun Layihə Hesabatı",
        "project": "Layihə", "report_id": "Hesabat ID", "status": "Status", "currency": "Valyuta",
        "confidence": "Panel etibarlılığı", "dashboard_type": "Panel növü", "dashboard_view": "İdarəetmə paneli",
        "selected_dashboard": "Seçilmiş analiz paneli",
        "report_follows": "Bu hesabat seçilmiş analiz panelinin məntiqinə əsasən hazırlanıb.",
        "metric": "Göstərici", "value": "Dəyər", "note": "Qeyd",
        "key_indicators": "Əsas göstəricilər", "planned_execution": "Plan üzrə icra", "actual_execution": "Faktiki icra",
        "progress_gap": "İcra fərqi", "delay": "Gecikmə", "total_cost": "Smeta / xərc hesablaması",
        "planned_cost": "Plan dəyəri", "actual_cost": "Faktiki dəyər", "cost_variance": "Xərc fərqi",
        "workforce": "İşçi sayı", "risk_score": "Risk balı", "risk_level": "Risk səviyyəsi",
        "baseline_finish": "Plan üzrə bitmə", "estimated_finish": "Proqnoz bitmə",
        "executive_summary": "Rəhbərlik xülasəsi", "recommended_actions": "Tövsiyə olunan tədbirlər",
        "no_action": "Mövcud məlumatlara əsasən tövsiyə olunan tədbir yaradılmadı.",
        "risk_register": "Risk reyestri", "risk": "Risk", "level": "Səviyyə", "reason": "Səbəb", "action": "Tədbir",
        "data_quality": "Məlumat keyfiyyəti qeydləri", "sheet_profiles": "Aşkarlanmış vərəq profilləri",
        "file": "Fayl", "sheet": "Vərəq", "type": "Növ", "rows": "Sətir", "not_available": "Mövcud deyil",
        "forecast": "Proqnoz", "unit": "Vahid", "panel": "Panel", "kpi": "KPI", "warning": "Xəbərdarlıq",
        "header_row": "Başlıq sətri", "mapped_columns": "Uyğunlaşdırılmış sütunlar", "signals": "Siqnallar", "days": "gün", "required": "tələb olunan"
    }
}

def _lang(lang: str | None) -> str:
    return "az" if str(lang or "en").lower().startswith("az") else "en"

def _label(key: str, lang: str = "en") -> str:
    return _LABELS.get(_lang(lang), _LABELS["en"]).get(key, key)

def _translate_na(value: Any, lang: str = "en") -> str:
    text = _text(value)
    return _label("not_available", lang) if text == "Not available" else text

def build_pdf_bytes(result: Dict[str, Any], lang: str = "en", paper: str = "a4") -> bytes:
    lang = _lang(lang)
    _register_unicode_fonts()

    dashboard = result["dashboard"]
    project = dashboard["project"]
    kpis = dashboard["kpis"]
    forecast = dashboard["forecast"]

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=_paper_size(paper),
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=f"DevBareun Report - {project.get('name', '')}",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "DevBareunTitle",
        parent=styles["Title"],
        fontName=_BOLD_FONT,
        fontSize=18,
        leading=23,
        alignment=TA_LEFT,
        spaceAfter=8,
        textColor=colors.HexColor("#0f172a"),
    )
    h2 = ParagraphStyle(
        "DevBareunH2",
        parent=styles["Heading2"],
        fontName=_BOLD_FONT,
        fontSize=12,
        leading=15,
        spaceBefore=10,
        spaceAfter=6,
        textColor=colors.HexColor("#0f172a"),
    )
    body = ParagraphStyle(
        "DevBareunBody",
        parent=styles["BodyText"],
        fontName=_REGULAR_FONT,
        fontSize=9.2,
        leading=12.6,
        spaceAfter=5,
        alignment=TA_LEFT,
    )
    small = ParagraphStyle(
        "DevBareunSmall",
        parent=body,
        fontSize=8,
        leading=10.8,
    )
    header = ParagraphStyle(
        "DevBareunTableHeader",
        parent=small,
        fontName=_BOLD_FONT,
        textColor=colors.white,
        alignment=TA_CENTER,
    )
    label = ParagraphStyle(
        "DevBareunLabel",
        parent=small,
        fontName=_BOLD_FONT,
    )

    story: List[Any] = []
    story.append(Paragraph(_label("report_title", lang), title_style))

    meta_data = [
        [_paragraph(_label("project", lang), label), _paragraph(project.get("name"), body)],
        [_paragraph(_label("report_id", lang), label), _paragraph(project.get("report_id"), body)],
        [_paragraph(_label("status", lang), label), _paragraph(project.get("status"), body)],
        [_paragraph(_label("currency", lang), label), _paragraph(project.get("currency"), body)],
        [_paragraph(_label("confidence", lang), label), _paragraph(f"{_text(project.get('confidence'))}/100", body)],
        [_paragraph(_label("dashboard_type", lang), label), _paragraph(project.get("dashboard_title") or project.get("analysis_type") or _label("dashboard_view", lang), body)],
    ]
    story.append(_styled_table(meta_data, [42 * mm, 126 * mm]))
    story.append(Spacer(1, 5))

    dashboard_sections = dashboard.get("dashboard_sections", {}) or {}
    primary_kpis = dashboard_sections.get("primary_kpis", []) or []
    story.append(Paragraph(project.get("dashboard_title") or dashboard_sections.get("title") or _label("dashboard_view", lang), h2))
    story.append(Paragraph(xml_escape(_text(project.get("dashboard_description") or dashboard_sections.get("description") or _label("report_follows", lang))), body))
    if primary_kpis:
        view_rows = [[Paragraph(_label("metric", lang), header), Paragraph(_label("value", lang), header), Paragraph(_label("note", lang), header)]]
        for item in primary_kpis:
            view_rows.append([
                _paragraph(item.get("label"), small),
                _paragraph(_format_dashboard_metric(item), small),
                _paragraph(item.get("note") or item.get("status") or "", small),
            ])
        view_table = Table(view_rows, colWidths=[58 * mm, 48 * mm, 62 * mm], repeatRows=1)
        view_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(view_table)

    premium = dashboard.get("premium_dashboard") or {}
    if premium:
        story.append(Paragraph("Full Project Control Premium Sections", h2))
        premium_rows = [[Paragraph("Section", header), Paragraph("Metric", header), Paragraph("Value", header)]]
        for row in _premium_section_rows(premium):
            premium_rows.append([
                _paragraph(row.get("section"), small),
                _paragraph(row.get("metric"), small),
                _paragraph(row.get("value"), small),
            ])
        premium_table = Table(premium_rows, colWidths=[42 * mm, 54 * mm, 72 * mm], repeatRows=1)
        premium_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(premium_table)

    story.append(Paragraph(_label("key_indicators", lang), h2))
    kpi_data = [
        [_paragraph(_label("planned_execution", lang), label), _paragraph(_format_percent(kpis.get("planned_execution"), lang), body)],
        [_paragraph(_label("actual_execution", lang), label), _paragraph(_format_percent(kpis.get("actual_execution"), lang), body)],
        [_paragraph(_label("progress_gap", lang), label), _paragraph(_format_percent(kpis.get("schedule_gap_percent"), lang), body)],
        [_paragraph(_label("delay", lang), label), _paragraph(_format_days(kpis.get("delay_days"), lang), body)],
        [_paragraph(_label("total_cost", lang), label), _paragraph(_format_money(kpis.get("total_cost"), project.get("currency"), lang), body)],
        [_paragraph(_label("planned_cost", lang), label), _paragraph(_format_money(kpis.get("planned_cost"), project.get("currency"), lang), body)],
        [_paragraph(_label("actual_cost", lang), label), _paragraph(_format_money(kpis.get("actual_cost"), project.get("currency"), lang), body)],
        [_paragraph(_label("cost_variance", lang), label), _paragraph(_format_percent(kpis.get("cost_variance_percent"), lang), body)],
        [_paragraph(_label("workforce", lang), label), _paragraph(_format_workforce(kpis.get("workforce_current"), kpis.get("workforce_required"), lang), body)],
        [_paragraph(_label("risk_score", lang), label), _paragraph(_format_risk(kpis.get("risk_score"), lang), body)],
        [_paragraph(_label("risk_level", lang), label), _paragraph(kpis.get("risk_level"), body)],
        [_paragraph(_label("baseline_finish", lang), label), _paragraph(forecast.get("baseline_finish"), body)],
        [_paragraph(_label("estimated_finish", lang), label), _paragraph(forecast.get("estimated_finish"), body)],
    ]
    story.append(_styled_table(kpi_data, [50 * mm, 118 * mm]))

    story.append(Paragraph(_label("executive_summary", lang), h2))
    story.append(Paragraph(xml_escape(_text(dashboard.get("executive_summary"))), body))

    story.append(Paragraph(_label("recommended_actions", lang), h2))
    actions = dashboard.get("recommended_actions", []) or []
    if actions:
        for idx, action in enumerate(actions, start=1):
            story.append(Paragraph(f"{idx}. {xml_escape(_text(action))}", body))
    else:
        story.append(Paragraph(_label("no_action", lang), body))

    story.append(Paragraph(_label("risk_register", lang), h2))
    risk_rows = [[
        Paragraph(_label("risk", lang), header),
        Paragraph(_label("level", lang), header),
        Paragraph(_label("reason", lang), header),
        Paragraph(_label("action", lang), header),
    ]]
    for row in dashboard.get("risk_register", []) or []:
        risk_rows.append([
            _paragraph(row.get("risk"), small),
            _paragraph(row.get("level"), small),
            _paragraph(row.get("reason"), small),
            _paragraph(row.get("action"), small),
        ])
    table = Table(risk_rows, colWidths=[34 * mm, 22 * mm, 56 * mm, 56 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("FONTNAME", (0, 0), (-1, 0), _BOLD_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(table)

    warnings = dashboard.get("data_quality", {}).get("warnings", []) or []
    if warnings:
        story.append(Paragraph(_label("data_quality", lang), h2))
        for warning in warnings:
            story.append(Paragraph(f"- {xml_escape(_text(warning))}", body))

    sheets = dashboard.get("data_quality", {}).get("sheet_profiles", []) or []
    if sheets:
        story.append(Paragraph(_label("sheet_profiles", lang), h2))
        sheet_rows = [[
            Paragraph(_label("file", lang), header),
            Paragraph(_label("sheet", lang), header),
            Paragraph(_label("type", lang), header),
            Paragraph(_label("confidence", lang), header),
            Paragraph(_label("rows", lang), header),
        ]]
        for sheet in sheets[:20]:
            sheet_rows.append([
                _paragraph(sheet.get("file_name"), small),
                _paragraph(sheet.get("sheet_name"), small),
                _paragraph(sheet.get("detected_type"), small),
                _paragraph(sheet.get("confidence"), small),
                _paragraph(sheet.get("row_count"), small),
            ])
        sheet_table = Table(sheet_rows, colWidths=[48 * mm, 44 * mm, 30 * mm, 25 * mm, 21 * mm], repeatRows=1)
        sheet_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(sheet_table)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def build_excel_bytes(result: Dict[str, Any], lang: str = "en") -> bytes:
    lang = _lang(lang)
    dashboard = result["dashboard"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard KPIs"
    ws.append([_label("report_title", lang)])
    ws.append([])
    ws.append([_label("project", lang), dashboard["project"].get("name")])
    ws.append([_label("report_id", lang), dashboard["project"].get("report_id")])
    ws.append(["Result ID", dashboard["project"].get("result_id") or dashboard["project"].get("report_id")])
    ws.append(["Project ID", result.get("project_id") or dashboard.get("project_id") or "—"])
    ws.append([_label("status", lang), dashboard["project"].get("status")])
    ws.append([_label("currency", lang), dashboard["project"].get("currency")])
    ws.append([_label("confidence", lang), dashboard["project"].get("confidence")])
    ws.append([])
    ws.append([_label("kpi", lang), _label("value", lang)])
    for key, value in dashboard["kpis"].items():
        ws.append([key, value if value is not None else _label("not_available", lang)])
    ws.append([])
    ws.append([_label("forecast", lang), _label("value", lang)])
    for key, value in dashboard["forecast"].items():
        ws.append([key, value if value is not None else _label("not_available", lang)])
    _style_sheet(ws)

    ws_view = wb.create_sheet(_label("dashboard_view", lang))
    ws_view.append([dashboard["project"].get("dashboard_title") or _label("dashboard_view", lang)])
    ws_view.append([dashboard["project"].get("dashboard_description") or _label("selected_dashboard", lang)])
    ws_view.append([])
    ws_view.append([_label("metric", lang), _label("value", lang), _label("unit", lang), _label("status", lang), _label("note", lang)])
    for item in (dashboard.get("dashboard_sections", {}) or {}).get("primary_kpis", []) or []:
        ws_view.append([
            item.get("label"),
            item.get("value") if item.get("value") is not None else "Not available",
            item.get("unit"),
            item.get("status"),
            item.get("note"),
        ])
    ws_view.append([])
    ws_view.append([_label("panel", lang), _label("metric", lang), _label("value", lang), _label("unit", lang)])
    for panel in (dashboard.get("dashboard_sections", {}) or {}).get("panels", []) or []:
        for item in panel.get("rows", []) or []:
            ws_view.append([
                panel.get("title"),
                item.get("label"),
                item.get("value") if item.get("value") is not None else "Not available",
                item.get("unit"),
            ])
    _style_sheet(ws_view)

    ws2 = wb.create_sheet(_label("risk_register", lang))
    ws2.append([_label("risk", lang), _label("level", lang), _label("reason", lang), _label("action", lang)])
    for row in dashboard.get("risk_register", []):
        ws2.append([row.get("risk"), row.get("level"), row.get("reason"), row.get("action")])
    _style_sheet(ws2)

    ws3 = wb.create_sheet(_label("recommended_actions", lang))
    ws3.append(["No", _label("recommended_actions", lang)])
    for idx, action in enumerate(dashboard.get("recommended_actions", []), 1):
        ws3.append([idx, action])
    _style_sheet(ws3)

    ws4 = wb.create_sheet(_label("sheet_profiles", lang))
    ws4.append([_label("file", lang), _label("sheet", lang), _label("type", lang), _label("confidence", lang), _label("header_row", lang), _label("mapped_columns", lang), _label("signals", lang), _label("rows", lang)])
    for sheet in dashboard.get("data_quality", {}).get("sheet_profiles", []):
        ws4.append([
            sheet.get("file_name"),
            sheet.get("sheet_name"),
            sheet.get("detected_type"),
            sheet.get("confidence"),
            sheet.get("header_row"),
            str(sheet.get("mapped_columns", {})),
            ", ".join(sheet.get("signals", [])),
            sheet.get("row_count"),
        ])
    _style_sheet(ws4)

    ws5 = wb.create_sheet(_label("data_quality", lang))
    ws5.append([_label("warning", lang)])
    for warning in dashboard.get("data_quality", {}).get("warnings", []):
        ws5.append([warning])
    _style_sheet(ws5)

    premium = dashboard.get("premium_dashboard") or {}
    if premium:
        ws_premium = wb.create_sheet("Premium Sections")
        ws_premium.append(["Section", "Metric", "Value"])
        for row in _premium_section_rows(premium):
            ws_premium.append([row.get("section"), row.get("metric"), row.get("value")])
        _style_sheet(ws_premium)

        ws_actions = wb.create_sheet("Recovery Actions")
        ws_actions.append(["Module", "Action", "Priority", "Status"])
        for row in premium.get("recovery_actions") or []:
            ws_actions.append([row.get("module"), row.get("action"), row.get("priority"), row.get("status")])
        _style_sheet(ws_actions)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _premium_section_rows(premium: Dict[str, Any]) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for section_name, section in [
        ("Executive Summary", premium.get("executive_summary") or {}),
        ("KPI Summary", premium.get("kpis") or {}),
        ("Schedule Analysis", premium.get("schedule_analysis") or {}),
        ("Cost & Payment Analysis", premium.get("cost_payment_analysis") or {}),
        ("Workforce Analysis", premium.get("workforce_analysis") or {}),
        ("Material Continuity", premium.get("material_continuity") or {}),
        ("Risk Register", premium.get("risk_register_analysis") or {}),
        ("Data Quality", premium.get("data_quality") or {}),
    ]:
        if not isinstance(section, dict):
            continue
        for key, value in section.items():
            if isinstance(value, (dict, list)):
                continue
            rows.append({"section": section_name, "metric": str(key).replace("_", " ").title(), "value": _text(value)})
    if not rows:
        rows.append({"section": "Full Project Control Premium", "metric": "Status", "value": "No premium section data available"})
    return rows[:80]


def _styled_table(rows: List[List[Any]], widths: List[float]) -> Table:
    table = Table(rows, colWidths=widths)
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f1f5f9")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return table


def _style_sheet(ws) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in ws[1]:
        cell.font = Font(bold=True, size=14)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        values = [cell.value for cell in row]
        if values and values[0] in {"KPI", "Forecast", "Risk", "No", "File", "Warning",
            "Metric", "Value", "Note", "Panel", "Sheet", "Type", "Rows",
            "Section", "Module", "Action", "Priority", "Status",
            "Proqnoz", "Fayl", "Xəbərdarlıq", "Göstərici",
            "Məlumat keyfiyyəti qeydləri", "Dəyər", "Qeyd", "Vərəq", "Növ",
            "Tövsiyə olunan tədbirlər", "Risk reyestri",
            "Aşkarlanmış vərəq profilləri", "Layihə"}:
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
    for column_cells in ws.columns:
        max_len = max(len(str(c.value or "")) for c in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(12, max_len + 2), 60)


def _format_dashboard_metric(item: Any) -> str:
    """Format a primary_kpi item dict {value, unit, label} for PDF display."""
    if not item or not isinstance(item, dict):
        return "Not available"
    value = item.get("value")
    unit = item.get("unit", "")
    if value is None:
        return "Not available"
    try:
        number = float(value)
        if unit in ("%",):
            return f"{number:,.2f}%"
        if unit:
            return f"{number:,.2f} {unit}"
        return f"{number:,.2f}"
    except Exception:
        return str(value)


def _format_percent(value: Any, lang: str = "en") -> str:
    return _label("not_available", lang) if value is None else f"{value}%"


def _format_money(value: Any, currency: Any = None, lang: str = "en") -> str:
    if value is None:
        return _label("not_available", lang)
    try:
        number = float(value)
        suffix = f" {currency}" if currency else ""
        return f"{number:,.2f}{suffix}"
    except Exception:
        return _text(value)


def _format_days(value: Any, lang: str = "en") -> str:
    if value is None:
        return _label("not_available", lang)
    suffix = _label("days", lang)
    return f"+{value} {suffix}" if value and value > 0 else f"{value} {suffix}"


def _format_risk(value: Any, lang: str = "en") -> str:
    return _label("not_available", lang) if value is None else f"{value}/100"


def _format_workforce(current: Any, required: Any, lang: str = "en") -> str:
    if current is None and required is None:
        return _label("not_available", lang)
    if required is None:
        return _text(current)
    req_label = _label("required", lang)
    return f"{_text(current)} / {req_label} {required}"
