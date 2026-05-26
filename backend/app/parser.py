from __future__ import annotations

import csv
import re
import xml.etree.ElementTree as ET
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from dateutil import parser as date_parser
from openpyxl import load_workbook

from .models import ParsedProjectData, SheetProfile
from .productivity import analyze_workforce_productivity

try:
    import pandas as pd
except Exception:  # pragma: no cover - pandas is optional for xls fallback
    pd = None


TYPE_KEYWORDS: Dict[str, Sequence[str]] = {
    "cost": (
        "smeta", "boq", "bill of quantity", "quantity", "qty", "miqdar", "həcim", "hecim",
        "vahid", "unit", "unit price", "qiymət", "qiymet", "məbləğ", "mebleg", "amount",
        "total", "cəmi", "cemi", "əDV", "edv", "vat", "contract value", "müqavilə dəyəri",
        "muqavile deyeri", "xərc", "xerc", "cost", "budget", "forecast overrun"
    ),
    "schedule": (
        "schedule", "qrafik", "baseline", "plan start", "plan finish", "planned start", "planned finish",
        "start", "finish", "duration", "müddət", "muddet", "calendar", "primavera", "xer",
        "wbs", "activity", "fəaliyyət", "fealiyyet", "critical", "completion date"
    ),
    "progress": (
        "progress", "execution", "icra", "faktiki", "actual", "planned", "plan üzrə", "plan uzre",
        "forma-2", "forma 2", "f-2", "f2", "hakediş", "hakedis", "payment certificate", "percent", "%",
        "physical progress", "done", "completed"
    ),
    "workforce": (
        "workforce", "labor", "labour", "manpower", "işçi", "isci", "fəhlə", "fehle", "usta",
        "crew", "brigade", "briqada", "worker", "required workforce", "cari işçi"
    ),
    "procurement": (
        "procurement", "supplier", "təchizat", "techizat", "delivery", "material", "purchase",
        "sifariş", "sifaris", "stock", "warehouse", "anbar", "lead time"
    ),
    "report": (
        "executive summary", "recommended actions", "risk score", "report id", "project report",
        "rəhbərlik xülasəsi", "rehberlik xulasesi", "tövsiyə", "tovsiye", "hesabat"
    ),
}

COLUMN_KEYWORDS: Dict[str, Sequence[str]] = {
    "planned_execution": ("planned execution", "plan üzrə icra", "plan uzre icra", "planned %", "plan %", "baseline %", "plan"),
    "actual_execution": ("actual execution", "faktiki icra", "actual %", "faktiki %", "icra faizi", "completed %", "progress %", "actual"),
    "delay_days": ("delay", "gecikmə", "gecikme", "delay days", "gün", "days late"),
    "cost_variance_percent": ("cost variance", "xərc fərqi", "xerc ferqi", "variance %", "budget variance", "overrun"),
    "workforce_current": ("current workforce", "cari işçi", "cari isci", "worker count", "manpower", "işçi sayı", "isci sayi", "current"),
    "workforce_required": ("required workforce", "tələb olunan", "teleb olunan", "required", "needed workforce"),
    "baseline_finish": ("baseline finish", "plan üzrə bitmə", "plan uzre bitme", "planned finish", "target finish"),
    "estimated_finish": ("estimated finish", "proqnoz", "forecast finish", "revised finish", "expected finish"),
    "planned_cost": ("planned cost", "baseline cost", "budget", "plan xərc", "plan xerc"),
    "actual_cost": ("actual cost", "faktiki xərc", "faktiki xerc", "spent", "burn"),
    "amount": ("amount", "məbləğ", "mebleg", "cəmi", "cemi", "total", "sum"),
    "quantity": ("quantity", "qty", "miqdar", "həcim", "hecim"),
    "unit_price": ("unit price", "vahid qiymət", "vahid qiymet"),
    "currency": ("currency", "valyuta"),
}

PROJECT_LABELS = (
    "project", "project name", "layihə", "layihe", "layihənin adı", "layihenin adi",
    "obyekt", "object", "müqavilə predmeti", "muqavile predmeti", "işlərin adı", "islerin adi"
)

AZ_TR_CHARS = set("əƏğĞıİöÖşŞüÜçÇ")
MAX_REASONABLE_MONEY = 1_000_000_000.0


class ConstructionFileParser:
    def __init__(self, analysis_type: str | None = None) -> None:
        self.analysis_type = (analysis_type or "all").lower()

    def parse_files(self, paths: Sequence[Path]) -> ParsedProjectData:
        parsed = ParsedProjectData()
        self._workbook_metric_evidence: List[Dict[str, Any]] = []
        self._dashboard_input_evidence: List[Dict[str, Any]] = []
        self._az_f2_special_results: List[Dict[str, Any]] = []
        all_text: List[str] = []
        all_rows: List[List[Any]] = []
        project_candidates: List[str] = []

        # Azerbaijan Nokopitelni/F-2 workbooks need a deterministic parser before
        # generic column inference. This protects the dashboard from reading line-item
        # quantities or concatenated cells as actual cost.
        for f2_path in [p for p in paths if p.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}]:
            try:
                f2_result = self._try_az_f2_smeta_parse(f2_path)
                if f2_result:
                    self._az_f2_special_results.append(f2_result)
            except Exception:
                # Keep upload resilient; generic parsing will still run and warnings are
                # handled by validation/preflight.
                pass

        for path in paths:
            suffix = path.suffix.lower()
            try:
                if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
                    file_rows, candidates, sheets = self._parse_xlsx(path)
                elif suffix == ".csv":
                    file_rows, candidates, sheets = self._parse_csv(path)
                elif suffix == ".xls":
                    file_rows, candidates, sheets = self._parse_xls(path)
                elif suffix == ".pdf":
                    file_rows, candidates, sheets = self._parse_pdf(path)
                elif suffix in {".jpg", ".jpeg", ".png", ".webp"}:
                    file_rows, candidates, sheets = self._parse_lightweight(path)
                elif suffix == ".xer":
                    file_rows, candidates, sheets = self._parse_xer(path)
                elif suffix == ".xml":
                    file_rows, candidates, sheets = self._parse_msproject_xml(path)
                else:
                    parsed.warnings.append(f"Unsupported file type skipped: {path.name}")
                    continue
            except Exception as exc:
                parsed.warnings.append(f"Could not parse {path.name}: {exc}")
                continue

            all_rows.extend(file_rows)
            all_text.extend(self._row_text(r) for r in file_rows)
            project_candidates.extend(candidates)
            parsed.sheets.extend(sheets)

        joined = "\n".join(all_text)
        parsed.language_hint = self._language_hint(joined)
        parsed.currency = self._detect_currency(joined, parsed.language_hint)
        parsed.project_name = self._choose_project_name(project_candidates, paths)

        self._extract_metrics(parsed, all_rows, joined)
        self._apply_az_f2_special_results(parsed)
        self._apply_dashboard_input_evidence(parsed)
        if self.analysis_type in {"all", "workforce"}:
            self._extract_workforce_productivity(parsed, all_rows)
        self._post_process(parsed)
        return parsed

    def _parse_xlsx(self, path: Path) -> Tuple[List[List[Any]], List[str], List[SheetProfile]]:
        wb = load_workbook(path, data_only=True, read_only=True)
        rows: List[List[Any]] = []
        candidates: List[str] = []
        sheets: List[SheetProfile] = []
        sheet_bundles: List[Tuple[str, List[List[Any]]]] = []
        for ws in wb.worksheets:
            sheet_rows: List[List[Any]] = []
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                values = [self._clean_cell(v) for v in row]
                if any(v not in (None, "") for v in values):
                    sheet_rows.append(values)
                if idx >= 2000:
                    break
            rows.extend(sheet_rows)
            sheet_bundles.append((ws.title, sheet_rows))
            dashboard_evidence = self._extract_full_dashboard_input_evidence(path.name, ws.title, sheet_rows)
            if dashboard_evidence:
                self._dashboard_input_evidence.append(dashboard_evidence)
            candidates += self._project_candidates_from_rows(sheet_rows[:25])
            if self._looks_like_project_name(ws.title):
                candidates.append(ws.title)
            sheets.append(self._profile_sheet(path.name, ws.title, sheet_rows))
        evidence = self._extract_workbook_metric_evidence(path.name, sheet_bundles)
        if evidence:
            self._workbook_metric_evidence.append(evidence)
        return rows, candidates, sheets

    def _parse_csv(self, path: Path) -> Tuple[List[List[Any]], List[str], List[SheetProfile]]:
        rows: List[List[Any]] = []
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
            except Exception:
                dialect = csv.excel
            reader = csv.reader(f, dialect)
            for idx, row in enumerate(reader):
                rows.append([self._clean_cell(v) for v in row])
                if idx >= 2000:
                    break
        return rows, self._project_candidates_from_rows(rows[:25]), [self._profile_sheet(path.name, path.stem, rows)]

    def _parse_xls(self, path: Path) -> Tuple[List[List[Any]], List[str], List[SheetProfile]]:
        if pd is None:
            raise RuntimeError("pandas/xlrd is required for .xls files")
        rows: List[List[Any]] = []
        candidates: List[str] = []
        sheets: List[SheetProfile] = []
        sheet_bundles: List[Tuple[str, List[List[Any]]]] = []
        xls = pd.ExcelFile(path)
        for sheet in xls.sheet_names:
            frame = xls.parse(sheet, header=None, nrows=2000)
            sheet_rows = [[self._clean_cell(v) for v in row] for row in frame.values.tolist()]
            sheet_rows = [r for r in sheet_rows if any(v not in (None, "") for v in r)]
            rows.extend(sheet_rows)
            sheet_bundles.append((sheet, sheet_rows))
            dashboard_evidence = self._extract_full_dashboard_input_evidence(path.name, sheet, sheet_rows)
            if dashboard_evidence:
                self._dashboard_input_evidence.append(dashboard_evidence)
            candidates += self._project_candidates_from_rows(sheet_rows[:25])
            if self._looks_like_project_name(sheet):
                candidates.append(sheet)
            sheets.append(self._profile_sheet(path.name, sheet, sheet_rows))
        evidence = self._extract_workbook_metric_evidence(path.name, sheet_bundles)
        if evidence:
            self._workbook_metric_evidence.append(evidence)
        return rows, candidates, sheets

    def _parse_pdf(self, path: Path) -> Tuple[List[List[Any]], List[str], List[SheetProfile]]:
        """Extract text from text-based PDFs when possible.

        Scanned PDFs/images are not OCR'd in this MVP; they remain supporting
        documents. Text-based PDF reports, BOQ summaries and payment summaries can
        still contribute labels/amounts to the generic parser.
        """
        rows: List[List[Any]] = []
        signals = ["pdf"]
        try:
            from pypdf import PdfReader  # type: ignore

            reader = PdfReader(str(path))
            for page in reader.pages[:20]:
                text = page.extract_text() or ""
                for line in text.splitlines():
                    clean = self._clean_cell(line)
                    if clean:
                        rows.append([clean])
                        if len(rows) >= 2000:
                            break
                if len(rows) >= 2000:
                    break
        except Exception:
            rows = []

        if not rows:
            rows = [[path.stem]]
            detected_type = "supporting_document"
            confidence = 35
            signals.append("pdf-supporting-only")
        else:
            detected_type = "report"
            confidence = 58
            signals.append("pdf-text-extracted")

        profile = SheetProfile(
            file_name=path.name,
            sheet_name=path.stem,
            detected_type=detected_type,
            confidence=confidence,
            signals=signals,
            row_count=len(rows),
        )
        return rows, self._project_candidates_from_rows(rows[:25]) or [path.stem], [profile]

    def _parse_xer(self, path: Path) -> Tuple[List[List[Any]], List[str], List[SheetProfile]]:
        """Basic Primavera P6 XER schedule extraction.

        Extracts TASK rows into a normalized table when the XER contains the
        standard %T / %F / %R blocks. This is a schedule beta parser: it reads
        activities and common dates but does not yet perform full critical path
        or predecessor network analysis.
        """
        rows: List[List[Any]] = []
        header = [
            "Activity ID", "Activity Name", "WBS", "Planned Start", "Planned Finish",
            "Actual Start", "Actual Finish", "Percent Complete", "Status"
        ]
        rows.append(header)
        try:
            content = path.read_text(encoding="utf-8", errors="ignore").splitlines()
        except Exception:
            content = []

        current_table = None
        fields: List[str] = []
        task_count = 0
        for line in content:
            if not line:
                continue
            parts = line.split("	")
            marker = parts[0].strip()
            if marker == "%T":
                current_table = parts[1].strip() if len(parts) > 1 else None
                fields = []
            elif marker == "%F" and current_table == "TASK":
                fields = parts[1:]
            elif marker == "%R" and current_table == "TASK" and fields:
                record = dict(zip(fields, parts[1:]))
                activity_id = record.get("task_code") or record.get("task_id") or record.get("id")
                name = record.get("task_name") or record.get("name")
                if not (activity_id or name):
                    continue
                rows.append([
                    activity_id,
                    name,
                    record.get("wbs_id") or record.get("projwbs_id") or "",
                    self._clean_xer_date(record.get("target_start_date") or record.get("early_start_date") or record.get("start_date")),
                    self._clean_xer_date(record.get("target_end_date") or record.get("early_end_date") or record.get("end_date")),
                    self._clean_xer_date(record.get("act_start_date")),
                    self._clean_xer_date(record.get("act_end_date")),
                    self._parse_number(record.get("phys_complete_pct") or record.get("complete_pct") or record.get("pct_complete")),
                    record.get("status_code") or record.get("task_type") or "",
                ])
                task_count += 1
                if task_count >= 5000:
                    break

        if task_count == 0:
            rows = [[path.stem]]
            confidence = 35
            signals = ["xer", "supporting-schedule-export"]
        else:
            confidence = 78
            signals = ["Primavera XER", "TASK table", "schedule activity extraction"]

        profile = SheetProfile(
            file_name=path.name,
            sheet_name="Primavera XER",
            detected_type="schedule",
            confidence=confidence,
            header_row=1 if task_count else None,
            mapped_columns={
                "baseline_finish": "E",
                "estimated_finish": "G",
                "actual_execution": "H",
            } if task_count else {},
            signals=signals,
            row_count=len(rows),
        )
        return rows, [path.stem], [profile]

    def _parse_msproject_xml(self, path: Path) -> Tuple[List[List[Any]], List[str], List[SheetProfile]]:
        """Basic MS Project XML task extraction.

        Reads common Task fields such as UID, ID, Name, Start, Finish,
        ActualStart, ActualFinish and PercentComplete.
        """
        rows: List[List[Any]] = [[
            "Activity ID", "Activity Name", "Planned Start", "Planned Finish",
            "Actual Start", "Actual Finish", "Percent Complete", "Critical"
        ]]
        task_count = 0
        try:
            root = ET.parse(path).getroot()
            for task in root.iter():
                if self._xml_local_name(task.tag) != "Task":
                    continue
                values: Dict[str, Any] = {}
                for child in list(task):
                    name = self._xml_local_name(child.tag)
                    if name in {"UID", "ID", "Name", "Start", "Finish", "ActualStart", "ActualFinish", "PercentComplete", "Critical"}:
                        values[name] = child.text
                task_name = values.get("Name")
                if not task_name:
                    continue
                activity_id = values.get("ID") or values.get("UID")
                rows.append([
                    activity_id,
                    task_name,
                    self._clean_xml_date(values.get("Start")),
                    self._clean_xml_date(values.get("Finish")),
                    self._clean_xml_date(values.get("ActualStart")),
                    self._clean_xml_date(values.get("ActualFinish")),
                    self._parse_number(values.get("PercentComplete")),
                    values.get("Critical"),
                ])
                task_count += 1
                if task_count >= 5000:
                    break
        except Exception:
            task_count = 0

        if task_count == 0:
            rows = [[path.stem]]
            confidence = 35
            signals = ["xml", "supporting-schedule-document"]
        else:
            confidence = 82
            signals = ["MS Project XML", "Task extraction", "schedule activity extraction"]

        profile = SheetProfile(
            file_name=path.name,
            sheet_name="MS Project XML",
            detected_type="schedule",
            confidence=confidence,
            header_row=1 if task_count else None,
            mapped_columns={
                "baseline_finish": "D",
                "estimated_finish": "F",
                "actual_execution": "G",
            } if task_count else {},
            signals=signals,
            row_count=len(rows),
        )
        return rows, [path.stem], [profile]

    def _parse_lightweight(self, path: Path) -> Tuple[List[List[Any]], List[str], List[SheetProfile]]:
        # Image files are accepted as supporting site evidence in this release.
        # Visual recognition/OCR is intentionally not assumed, so no progress is
        # invented from photos.
        rows = [[path.stem]]
        profile = SheetProfile(
            file_name=path.name,
            sheet_name=path.stem,
            detected_type="site_image_supporting",
            confidence=35,
            signals=["image-supporting-evidence"],
            row_count=1,
        )
        return rows, [path.stem], [profile]

    def _clean_xer_date(self, value: Any) -> Optional[str]:
        if not value:
            return None
        text = str(value).strip()
        # Primavera often stores dates like 2026-01-15 08:00 or 2026-01-15T08:00
        parsed = self._parse_date(text)
        return parsed.isoformat() if parsed else text[:10]

    def _clean_xml_date(self, value: Any) -> Optional[str]:
        if not value:
            return None
        text = str(value).strip()
        parsed = self._parse_date(text)
        return parsed.isoformat() if parsed else text[:10]

    def _xml_local_name(self, tag: str) -> str:
        return str(tag).split('}', 1)[-1] if '}' in str(tag) else str(tag)

    def _profile_sheet(self, file_name: str, sheet_name: str, rows: Sequence[Sequence[Any]]) -> SheetProfile:
        text = "\n".join(self._row_text(r) for r in rows[:200]) + " " + sheet_name
        norm = self._norm(text)
        if self._is_nokopitelni_sheet(sheet_name):
            header_idx, mapped = self._detect_header_and_columns(rows)
            return SheetProfile(
                file_name=file_name,
                sheet_name=sheet_name,
                detected_type="cost",
                confidence=94 if mapped else 88,
                header_row=header_idx,
                mapped_columns=mapped,
                signals=["Nokopitelni", "smeta total", "cumulative estimate"],
                row_count=len(rows),
            )
        if self._is_f2_sheet(sheet_name, rows):
            header_idx, mapped = self._detect_header_and_columns(rows)
            return SheetProfile(
                file_name=file_name,
                sheet_name=sheet_name,
                detected_type="progress",
                confidence=90 if mapped else 82,
                header_row=header_idx,
                mapped_columns=mapped,
                signals=["F-2", "completed amount", "progress certificate"],
                row_count=len(rows),
            )
        scores: Dict[str, int] = {}
        signals: List[str] = []
        for dtype, keywords in TYPE_KEYWORDS.items():
            score = 0
            for kw in keywords:
                if self._norm(kw) in norm:
                    score += 1
                    if len(signals) < 12:
                        signals.append(kw)
            scores[dtype] = score
        detected = max(scores, key=scores.get) if scores else "unknown"
        top_score = scores.get(detected, 0)
        if top_score == 0:
            detected = "unknown"
            confidence = 20 if rows else 0
        else:
            confidence = min(95, 35 + top_score * 10)
        header_idx, mapped = self._detect_header_and_columns(rows)
        if mapped:
            confidence = min(98, confidence + min(20, len(mapped) * 4))
        detected, confidence, signals = self._apply_analysis_priority(detected, confidence, signals, mapped, sheet_name)
        return SheetProfile(
            file_name=file_name,
            sheet_name=sheet_name,
            detected_type=detected,
            confidence=confidence,
            header_row=header_idx,
            mapped_columns=mapped,
            signals=list(dict.fromkeys(signals)),
            row_count=len(rows),
        )


    def _apply_analysis_priority(self, detected: str, confidence: int, signals: List[str], mapped: Dict[str, str], sheet_name: str) -> Tuple[str, int, List[str]]:
        focus = getattr(self, "analysis_type", "all")
        if focus in ("all", ""):
            return detected, confidence, signals
        mapped_keys = set(mapped or {})
        norm_sheet = self._norm(sheet_name)
        rules = {
            "cost": {"types": {"cost", "progress"}, "columns": {"amount", "planned_cost", "actual_cost", "quantity", "unit_price"}},
            "progress": {"types": {"progress", "cost"}, "columns": {"planned_execution", "actual_execution", "actual_cost", "amount"}},
            "schedule": {"types": {"schedule"}, "columns": {"baseline_finish", "estimated_finish", "planned_execution", "actual_execution"}},
            "workforce": {"types": {"workforce"}, "columns": {"workforce_current", "workforce_required"}},
        }
        rule = rules.get(focus)
        if not rule:
            return detected, confidence, signals
        if detected in rule["types"] or mapped_keys & rule["columns"]:
            confidence = min(98, confidence + 8)
            signals = list(dict.fromkeys(list(signals) + [f"focus:{focus}"]))
        if focus == "cost" and any(t in norm_sheet for t in ("smeta", "boq", "nokop", "nakop")):
            detected = "cost"
            confidence = min(98, confidence + 10)
        if focus == "progress" and any(t in norm_sheet for t in ("f 2", "f-2", "forma", "progress", "icra")):
            detected = "progress"
            confidence = min(98, confidence + 10)
        return detected, confidence, signals

    def _detect_header_and_columns(self, rows: Sequence[Sequence[Any]]) -> Tuple[Optional[int], Dict[str, str]]:
        best_idx: Optional[int] = None
        best_score = 0
        best_map: Dict[str, str] = {}
        for idx, row in enumerate(rows[:40], start=1):
            mapped = self._map_columns(row)
            score = len(mapped) * 4 + sum(1 for v in row if v not in (None, ""))
            if score > best_score and len(mapped) >= 1:
                best_idx = idx
                best_score = score
                best_map = mapped
        return best_idx, best_map

    def _map_columns(self, header_row: Sequence[Any]) -> Dict[str, str]:
        mapped: Dict[str, str] = {}
        for col_index, value in enumerate(header_row):
            text = self._norm(str(value or ""))
            if not text:
                continue
            for standard, keywords in COLUMN_KEYWORDS.items():
                if any(self._norm(k) in text for k in keywords):
                    mapped.setdefault(standard, self._excel_column_name(col_index + 1))
        return mapped

    def _extract_metrics(self, parsed: ParsedProjectData, rows: Sequence[Sequence[Any]], text: str) -> None:
        parsed.planned_execution = self._extract_percent_by_labels(text, ("planned execution", "plan üzrə icra", "plan uzre icra", "plan %", "planned"))
        parsed.actual_execution = self._extract_percent_by_labels(text, ("actual execution", "faktiki icra", "actual %", "faktiki %", "icra səviyyəsi", "icra seviyesi"))
        parsed.cost_variance_percent = self._extract_percent_by_labels(text, ("cost variance", "xərc fərqi", "xerc ferqi", "variance", "overrun"))
        parsed.delay_days = self._extract_int_by_labels(text, ("delay", "gecikmə", "gecikme", "delay impact"))
        parsed.workforce_current = self._extract_int_by_labels(text, ("current workforce", "cari işçi", "cari isci", "işçi sayı", "isci sayi", "workforce"))
        parsed.workforce_required = self._extract_int_by_labels(text, ("required workforce", "tələb", "teleb", "required"))
        parsed.baseline_finish = self._extract_date_by_labels(text, ("baseline finish", "planned finish", "target finish", "plan üzrə bitmə", "plan uzre bitme"))
        parsed.estimated_finish = self._extract_date_by_labels(text, ("estimated finish", "forecast finish", "revised finish", "expected finish", "proqnoz"))
        parsed.planned_cost = self._extract_money_by_labels(text, ("planned cost", "baseline cost", "budget", "müqavilə dəyəri", "muqavile deyeri"))
        parsed.actual_cost = self._extract_money_by_labels(text, ("actual cost", "faktiki xərc", "faktiki xerc", "spent"))

        self._extract_from_tables(parsed, rows)
        self._apply_workbook_metric_evidence(parsed)

    def _extract_from_tables(self, parsed: ParsedProjectData, rows: Sequence[Sequence[Any]]) -> None:
        for i, row in enumerate(rows):
            label_text = self._norm(" ".join(str(v or "") for v in row[:4]))
            row_values = list(row)
            if parsed.planned_execution is None and self._has_any(label_text, ("planned execution", "plan uzre icra", "plan üzrə icra", "plan")):
                parsed.planned_execution = self._first_percent(row_values)
            if parsed.actual_execution is None and self._has_any(label_text, ("actual execution", "faktiki icra", "actual", "faktiki")):
                parsed.actual_execution = self._first_percent(row_values)
            if parsed.cost_variance_percent is None and self._has_any(label_text, ("cost variance", "xerc ferqi", "xərc fərqi", "variance", "overrun")):
                parsed.cost_variance_percent = self._first_percent(row_values)
            if parsed.delay_days is None and self._has_any(label_text, ("delay", "gecikme", "gecikmə")):
                parsed.delay_days = self._first_int(row_values)
            if parsed.baseline_finish is None and self._has_any(label_text, ("baseline finish", "planned finish", "target finish")):
                parsed.baseline_finish = self._first_date(row_values)
            if parsed.estimated_finish is None and self._has_any(label_text, ("estimated finish", "forecast finish", "revised finish")):
                parsed.estimated_finish = self._first_date(row_values)

            # Header + following rows pattern: inspect mapped numeric columns only
            # when the row looks like a real table header. KPI/value rows such as
            # "Cost | Smeta total | 3,139,625" can contain header-like words,
            # but treating them as headers pollutes totals with unrelated rows.
            mapped = self._map_columns(row)
            if mapped and i + 1 < len(rows) and self._looks_like_header_row(row, mapped):
                data_window = rows[i + 1 : min(i + 60, len(rows))]
                self._extract_from_mapped_columns(parsed, row, data_window, mapped)

    def _looks_like_header_row(self, row: Sequence[Any], mapped: Dict[str, str]) -> bool:
        non_empty = [v for v in row if v not in (None, "")]
        if not non_empty:
            return False
        numeric_or_date = 0
        for value in non_empty:
            if self._parse_date(value) or self._parse_number(value) is not None:
                numeric_or_date += 1
        # Most construction table headers are purely textual. If the row already
        # contains project values, amounts or dates, it is probably a data/KPI row.
        if numeric_or_date > 0:
            return False
        # Avoid weak one-column matches caused by broad words like "plan",
        # "actual" or "total" unless the sheet row clearly has header density.
        if len(mapped) < 2 and len(non_empty) < 4:
            return False
        return True

    def _extract_from_mapped_columns(
        self,
        parsed: ParsedProjectData,
        header: Sequence[Any],
        data_rows: Sequence[Sequence[Any]],
        mapped: Dict[str, str],
    ) -> None:
        def idx_for(col_name: str) -> Optional[int]:
            for idx in range(len(header)):
                if self._excel_column_name(idx + 1) == col_name:
                    return idx
            return None

        def collect_numbers(field: str, percent: bool = False) -> List[float]:
            idx = idx_for(mapped[field]) if field in mapped else None
            if idx is None:
                return []
            values: List[float] = []
            for row in data_rows:
                if idx < len(row):
                    val = self._parse_percent(row[idx]) if percent else self._parse_number(row[idx])
                    if val is not None:
                        values.append(val)
            return values

        if parsed.planned_execution is None and "planned_execution" in mapped:
            vals = collect_numbers("planned_execution", percent=True)
            if vals:
                parsed.planned_execution = max(vals) if max(vals) <= 100 else None
        if parsed.actual_execution is None and "actual_execution" in mapped:
            vals = collect_numbers("actual_execution", percent=True)
            if vals:
                parsed.actual_execution = max(vals) if max(vals) <= 100 else None
        if parsed.workforce_current is None and "workforce_current" in mapped:
            vals = collect_numbers("workforce_current")
            if vals:
                parsed.workforce_current = int(max(vals))
        if parsed.workforce_required is None and "workforce_required" in mapped:
            vals = collect_numbers("workforce_required")
            if vals:
                parsed.workforce_required = int(max(vals))
        if parsed.planned_cost is None and "planned_cost" in mapped:
            vals = [v for v in collect_numbers("planned_cost") if v > 0]
            if vals:
                parsed.planned_cost = sum(vals)
        if parsed.actual_cost is None and "actual_cost" in mapped:
            vals = [v for v in collect_numbers("actual_cost") if v > 0]
            if vals:
                parsed.actual_cost = sum(vals)
        if parsed.total_cost is None and "amount" in mapped:
            vals = [v for v in collect_numbers("amount") if v > 0]
            if vals:
                parsed.total_cost = sum(vals)

    def _apply_workbook_metric_evidence(self, parsed: ParsedProjectData) -> None:
        evidence_items = getattr(self, "_workbook_metric_evidence", [])
        if not evidence_items:
            return

        evidence_items = [e for e in evidence_items if e]
        if not evidence_items:
            return

        best = max(evidence_items, key=lambda e: float(e.get("smeta_total") or 0))
        smeta_total = best.get("smeta_total")
        f2_completed = best.get("f2_completed_amount")

        if smeta_total:
            smeta_total = float(smeta_total)
            parsed.total_cost = smeta_total
            if parsed.planned_cost is None:
                parsed.planned_cost = smeta_total
            parsed.evidence["smeta_total_source"] = best.get("smeta_total_source")

            # Safety rule: never let generic mapped-column extraction create absurd
            # actual cost values. If the detected actual cost is larger than 120% of
            # the smeta baseline, it needs user mapping confirmation instead of being
            # used in dashboard/PDF risk calculations.
            if parsed.actual_cost is not None and float(parsed.actual_cost) > smeta_total * 1.2:
                parsed.evidence["rejected_actual_cost"] = parsed.actual_cost
                parsed.warnings.append(
                    "Actual completed amount was rejected because it exceeded 120% of the smeta total. Please confirm the correct F-2 total or actual cost column."
                )
                parsed.actual_cost = None
                parsed.cost_variance_percent = None

        if smeta_total and f2_completed and smeta_total > 0:
            f2_completed = float(f2_completed)
            if 0 < f2_completed <= smeta_total * 1.2:
                execution = round((f2_completed / smeta_total) * 100, 2)
                parsed.actual_cost = f2_completed
                parsed.actual_execution = self._clamp(execution, 0, 100)
                parsed.evidence["actual_execution_source"] = "F-2 completed amount / nokopitelni smeta total"
                parsed.evidence["f2_completed_amount"] = round(f2_completed, 2)
                parsed.evidence["f2_sheets"] = best.get("f2_sheets", [])
                parsed.evidence["f2_execution_percent_raw"] = execution
                parsed.evidence["calculation_note"] = f"Actual execution calculated from validated F-2 certificates: {execution:g}% of smeta total."
            else:
                parsed.evidence["rejected_f2_completed_amount"] = f2_completed
                parsed.warnings.append(
                    "F-2 completed amount could not be safely confirmed because it exceeded 120% of the smeta total. Manual confirmation is required."
                )

    def _extract_full_dashboard_input_evidence(
        self,
        file_name: str,
        sheet_name: str,
        rows: Sequence[Sequence[Any]],
    ) -> Dict[str, Any]:
        """Read the official DevBareun Full_Dashboard_Input sheet deterministically.

        This sheet is generated by DevBareun templates and already separates
        planned, actual and variance columns. It must outrank generic table
        extraction so negative variance values are not misread as workforce
        counts or cost KPIs.
        """
        if self._norm(sheet_name) != "full dashboard input":
            return {}
        if not rows:
            return {}

        header_i: Optional[int] = None
        mapping: Dict[str, int] = {}
        required = {"category", "kpi", "planned", "actual"}
        for idx, row in enumerate(rows[:20]):
            candidate: Dict[str, int] = {}
            for col, value in enumerate(row):
                norm = self._norm(value)
                if norm in {"category", "kateqoriya"}:
                    candidate["category"] = col
                elif norm == "kpi" or "gosterici" in norm:
                    candidate["kpi"] = col
                elif norm in {"planned", "plan", "planlasdirilan"}:
                    candidate["planned"] = col
                elif norm in {"actual", "faktiki"}:
                    candidate["actual"] = col
                elif norm in {"variance", "ferq", "xerc ferqi"}:
                    candidate["variance"] = col
                elif norm == "date" or "tarix" in norm:
                    candidate["date"] = col
            if required.issubset(candidate):
                header_i = idx
                mapping = candidate
                break

        if header_i is None:
            return {}

        evidence: Dict[str, Any] = {
            "file": file_name,
            "sheet": sheet_name,
            "source": "Full_Dashboard_Input",
            "metrics": {},
        }

        def cell(row: Sequence[Any], key: str) -> Any:
            col = mapping.get(key)
            return row[col] if col is not None and col < len(row) else None

        for row in rows[header_i + 1: header_i + 80]:
            category = self._norm(cell(row, "category"))
            kpi = self._norm(cell(row, "kpi"))
            planned = self._parse_number(cell(row, "planned"))
            actual = self._parse_number(cell(row, "actual"))
            variance = self._parse_number(cell(row, "variance"))
            if not category and not kpi:
                continue

            if "cost" in category or "xerc" in category or "smeta" in kpi:
                if any(token in kpi for token in ("smeta", "estimate", "contract", "baseline", "total")) and planned and planned > 0:
                    evidence["metrics"]["total_cost"] = float(planned)
                    evidence["metrics"]["planned_cost"] = float(planned)
                if any(token in kpi for token in ("actual completed cost", "actual cost", "completed cost", "faktiki")) and actual and actual > 0:
                    evidence["metrics"]["actual_cost"] = float(actual)

            if "progress" in category or "icra" in category or "execution" in kpi:
                if planned is not None and 0 <= planned <= 100:
                    evidence["metrics"]["planned_execution"] = float(planned)
                if actual is not None and 0 <= actual <= 100:
                    evidence["metrics"]["actual_execution"] = float(actual)

            if "schedule" in category or "delay" in kpi or "gecik" in kpi:
                delay = actual if actual is not None else variance
                if delay is not None:
                    evidence["metrics"]["delay_days"] = max(0, int(round(delay)))

            if "workforce" in category or "worker" in kpi or "isci" in kpi:
                if planned is not None and planned >= 0:
                    evidence["metrics"]["workforce_required"] = int(round(planned))
                if actual is not None and actual >= 0:
                    evidence["metrics"]["workforce_current"] = int(round(actual))

        return evidence if evidence.get("metrics") else {}


    def _apply_dashboard_input_evidence(self, parsed: ParsedProjectData) -> None:
        evidence_items = getattr(self, "_dashboard_input_evidence", [])
        if not evidence_items:
            return

        merged: Dict[str, Any] = {}
        sources: List[Dict[str, Any]] = []
        for item in evidence_items:
            metrics = item.get("metrics") or {}
            if not metrics:
                continue
            merged.update(metrics)
            sources.append({"file": item.get("file"), "sheet": item.get("sheet"), "metrics": sorted(metrics.keys())})

        if not merged:
            return

        for field, value in merged.items():
            setattr(parsed, field, value)

        # Cost pair from the official template must own the variance formula.
        if parsed.planned_cost and parsed.actual_cost is not None:
            parsed.cost_variance_percent = round(((float(parsed.actual_cost) - float(parsed.planned_cost)) / float(parsed.planned_cost)) * 100, 2)
            parsed.evidence["authoritative_cost_pair"] = "Full_Dashboard_Input planned/actual columns"

        parsed.evidence["full_dashboard_input"] = sources
        parsed.evidence["template_priority_applied"] = True


    def _extract_workbook_metric_evidence(
        self,
        file_name: str,
        sheet_bundles: Sequence[Tuple[str, Sequence[Sequence[Any]]]],
    ) -> Dict[str, Any]:
        if not sheet_bundles:
            return {}

        smeta_candidates: List[Dict[str, Any]] = []
        f2_sheets: List[Dict[str, Any]] = []
        f2_completed_total = 0.0

        for sheet_name, rows in sheet_bundles:
            is_f2 = self._is_f2_sheet(sheet_name, rows)
            if is_f2:
                f2_amount, f2_source = self._find_sheet_total(rows, purpose="f2")
                if f2_amount is None and self._mapped_amount_column_is_period_amount(rows):
                    f2_amount, f2_source = self._sum_mapped_amount_column(rows, sheet_name)
                if f2_amount is None:
                    f2_amount, f2_source = self._last_mapped_amount_candidate(rows, sheet_name)
                if f2_amount is None:
                    f2_amount, f2_source = self._sum_mapped_amount_column(rows, sheet_name)
                if f2_amount is not None and f2_amount > 0:
                    f2_completed_total += float(f2_amount)
                    f2_sheets.append({
                        "sheet": sheet_name,
                        "amount": round(float(f2_amount), 2),
                        "source": f2_source,
                    })
                continue

            smeta_total, source = self._find_sheet_total(rows, purpose="smeta")
            if smeta_total is not None and smeta_total > 0:
                score = 0
                raw_name = str(sheet_name or "").lower()
                norm_name = self._norm(sheet_name)
                if self._is_nokopitelni_sheet(sheet_name):
                    score += 80
                if "smeta uzre yekun" in self._norm(source or ""):
                    score += 60
                if "yekun cem" in self._norm(source or ""):
                    score += 35
                if any(k in raw_name for k in ("nakop", "nokop", "накоп")) or "nokop" in norm_name:
                    score += 30
                score += min(20, int(float(smeta_total) / 100000))
                smeta_candidates.append({
                    "sheet": sheet_name,
                    "amount": round(float(smeta_total), 2),
                    "source": source,
                    "score": score,
                })

        best_smeta = None
        if smeta_candidates:
            smeta_candidates.sort(key=lambda x: (x["score"], x["amount"]), reverse=True)
            best_smeta = smeta_candidates[0]

        evidence: Dict[str, Any] = {"file_name": file_name}
        if best_smeta:
            evidence["smeta_total"] = best_smeta["amount"]
            evidence["smeta_total_source"] = {
                "file": file_name,
                "sheet": best_smeta["sheet"],
                "row": best_smeta["source"],
            }
            # Validate F-2 sheet totals against the selected baseline. This prevents
            # concatenated text/numeric cells from becoming impossible actual-cost values.
            upper = float(best_smeta["amount"]) * 1.2
            valid_f2_sheets = [x for x in f2_sheets if 0 < float(x.get("amount") or 0) <= upper]
            rejected_f2_sheets = [x for x in f2_sheets if float(x.get("amount") or 0) > upper]
            if valid_f2_sheets:
                valid_total = sum(float(x["amount"]) for x in valid_f2_sheets)
                if valid_total <= upper:
                    evidence["f2_completed_amount"] = round(valid_total, 2)
                    evidence["f2_sheets"] = valid_f2_sheets
                else:
                    evidence["f2_rejected_reason"] = "sum_of_f2_sheets_exceeds_120_percent_of_smeta"
                    evidence["f2_sheets_needing_confirmation"] = valid_f2_sheets
            if rejected_f2_sheets:
                evidence["rejected_f2_sheets"] = rejected_f2_sheets
        elif f2_sheets:
            # Without a baseline, keep the evidence for review but do not calculate progress.
            evidence["f2_sheets_needing_confirmation"] = f2_sheets
        return evidence

    def _is_f2_sheet(self, sheet_name: str, rows: Sequence[Sequence[Any]] | None = None) -> bool:
        # Nokopitelni / nakopitelni sheets are cumulative estimate sheets. They may contain
        # F-2 references in the header/body, but they must be used as the denominator
        # source, not as payment-certificate sheets.
        if self._is_nokopitelni_sheet(sheet_name):
            return False
        raw_name = str(sheet_name or "")
        norm_name = self._norm(raw_name).replace("-", " ")
        if re.search(r"\bf\s*2\b", norm_name) or "forma 2" in norm_name or "hak" in norm_name:
            return True
        if rows:
            sample = " ".join(self._row_text(r) for r in list(rows)[:12])
            norm_sample = self._norm(sample).replace("-", " ")
            if re.search(r"\bf\s*2\b", norm_sample) or "forma 2" in norm_sample or "hak" in norm_sample:
                return True
        return False

    def _is_nokopitelni_sheet(self, sheet_name: str) -> bool:
        raw = str(sheet_name or "").lower()
        norm = self._norm(sheet_name)
        return any(token in raw for token in ("nokop", "nakop", "накоп")) or any(token in norm for token in ("nokop", "nakop"))

    def _find_sheet_total(self, rows: Sequence[Sequence[Any]], purpose: str) -> Tuple[Optional[float], Optional[str]]:
        candidates: List[Tuple[int, float, str]] = []
        for row in rows:
            row_text = self._row_text(row)
            norm = self._norm(row_text)
            label_score = self._total_label_score(norm, purpose)
            if label_score <= 0:
                continue
            amount = self._best_amount_from_row(row)
            if amount is None:
                continue
            # Ignore tiny quantities and percentages; totals in construction estimates are usually > 1,000 AZN.
            if amount < 1000:
                continue
            candidates.append((label_score, float(amount), row_text[:240]))
        if not candidates:
            return None, None
        candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
        return candidates[0][1], candidates[0][2]

    def _total_label_score(self, norm: str, purpose: str) -> int:
        score = 0
        if "smeta uzre yekun" in norm or "smeta uzre" in norm:
            score += 100
        if purpose == "smeta" and any(token in norm for token in (
            "smeta total", "estimate total", "boq total", "contract total",
            "cost estimate total", "total estimate", "grand total"
        )):
            score += 90
        if "yekun cem" in norm or "yekun cemi" in norm or "umumi yekun" in norm:
            score += 80
        elif "yekun" in norm and "cem" in norm:
            score += 70
        elif re.search(r"\bcemi\b|\bcem\b", norm):
            score += 35
        if purpose == "f2" and any(token in norm for token in (
            "f 2", "forma 2", "yerine yetirilmis", "icra",
            "progress payment", "completed amount", "cumulative amount", "this period amount"
        )):
            score += 20
        if any(bad in norm for bad in ("edv", "vat", "2%", "faiz", "percent")) and score < 80:
            score -= 25
        return score

    def _best_amount_from_row(self, row: Sequence[Any]) -> Optional[float]:
        amounts: List[float] = []
        for value in row:
            if value in (None, ""):
                continue
            text = str(value)
            if "%" in text:
                continue
            if self._parse_date(value):
                continue
            number = self._parse_number(value)
            if number is None:
                continue
            if 1000 <= abs(number) <= MAX_REASONABLE_MONEY:
                amounts.append(abs(float(number)))
        if not amounts:
            return None
        return max(amounts)

    def _last_mapped_amount_candidate(self, rows: Sequence[Sequence[Any]], sheet_name: str) -> Tuple[Optional[float], Optional[str]]:
        """Return a likely F-2 sheet total from the mapped amount column.

        Many local F-2/Nokopitelni workbooks contain the certified total as the
        last large numeric value in the amount column, while individual line items
        appear above it. This avoids both false zero extraction and uncontrolled
        double counting. If a labelled total row exists, _find_sheet_total handles it
        first; this is a fallback for poorly labelled sheets.
        """
        header_idx, mapped = self._detect_header_and_columns(rows)
        col = mapped.get("amount") if mapped else None
        if not col or header_idx is None:
            return None, None
        col_idx = self._column_index(col)
        candidates: List[Tuple[int, float, str]] = []
        for pos, row in enumerate(rows[header_idx:], start=header_idx + 1):
            if col_idx >= len(row):
                continue
            number = self._parse_number(row[col_idx])
            if number is None or number < 1000:
                continue
            row_text = self._row_text(row)[:180]
            norm = self._norm(row_text)
            # Prefer visible total/subtotal rows near the bottom.
            score = pos
            if self._total_label_score(norm, "f2") > 0:
                score += 10000
            if any(token in norm for token in ("yekun", "cem", "cemi", "smeta", "icra", "yerine yetirilmis")):
                score += 3000
            candidates.append((score, float(number), row_text))
        if not candidates:
            return None, None
        candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
        return candidates[0][1], f"last/labelled mapped amount column {col} in {sheet_name}: {candidates[0][2]}"

    def _sum_mapped_amount_column(self, rows: Sequence[Sequence[Any]], sheet_name: str) -> Tuple[Optional[float], Optional[str]]:
        header_idx, mapped = self._detect_header_and_columns(rows)
        col = mapped.get("amount") if mapped else None
        if not col or header_idx is None:
            return None, None
        col_idx = self._column_index(col)
        total = 0.0
        count = 0
        # header_idx is 1-based.
        for row in rows[header_idx:]:
            row_text = self._norm(self._row_text(row))
            if self._total_label_score(row_text, "f2") > 0:
                continue
            if col_idx < len(row):
                number = self._parse_number(row[col_idx])
                if number is not None and number > 0:
                    total += float(number)
                    count += 1
        if total > 0 and count > 0:
            return total, f"summed mapped amount column {col} in {sheet_name} ({count} rows)"
        return None, None

    def _mapped_amount_column_is_period_amount(self, rows: Sequence[Sequence[Any]]) -> bool:
        header_idx, mapped = self._detect_header_and_columns(rows)
        col = mapped.get("amount") if mapped else None
        if not col or header_idx is None or header_idx < 1 or header_idx > len(rows):
            return False
        col_idx = self._column_index(col)
        header = rows[header_idx - 1]
        header_text = self._norm(str(header[col_idx] if col_idx < len(header) else ""))
        return any(token in header_text for token in (
            "this period amount", "period amount", "monthly amount",
            "current amount", "current period", "bu dovr", "cari dovr",
            "bu ay", "ayliq", "ayliq mebleg"
        ))

    def _column_index(self, col_name: str) -> int:
        value = 0
        for ch in str(col_name).upper():
            if "A" <= ch <= "Z":
                value = value * 26 + (ord(ch) - ord("A") + 1)
        return max(0, value - 1)

    # ══════════════════════════════════════════════════════════════════════
    # AZERBAIJANI NOKOPITELNI / F-2 PARSER
    # Uses the uploaded deterministic F-2 logic as a focused extractor for
    # local smeta/hakediş workbooks. Generic parsing still runs for all other
    # workbook types.
    # ══════════════════════════════════════════════════════════════════════

    _AZ_F2_COL_PATTERNS = (
        r"f-?\s*2\b", r"forma\s*-?\s*2", r"hakedis", r"hakedis", r"hakedi[sş]",
        r"akt\b", r"payment\s*cert", r"tamamlanan", r"icra",
    )
    _AZ_REMAINING_COL_PATTERNS = (r"qaliq", r"qalıq", r"remaining", r"balans", r"qalib", r"qalıb")
    _AZ_SMETA_TOTAL_PATTERNS = (
        r"smeta\s+uzre\s+yekun",
        r"smeta\s+uzr[e3]\s+yekun",
        r"smeta\s+[a-z]+\s+yekun",
        r"smeta\s+yekun",
        r"yekun\s+cem",
        r"yekun\s+cemi",
        r"umumi\s+yekun",
        r"contract\s+total",
        r"muqavile\s+cemi",
    )
    _AZ_SMETA_SHEET_NAMES = (
        "nokopitelni", "nakopitelni", "smeta", "boq", "muqavile", "müqavilə",
        "contract", "budget", "baseline",
    )
    _AZ_F2_SHEET_PATTERN = re.compile(r"f-?\s*2|forma\s*-?\s*2|hakedis|hakediş", re.IGNORECASE)

    def _apply_az_f2_special_results(self, parsed: ParsedProjectData) -> None:
        results = getattr(self, "_az_f2_special_results", []) or []
        if not results:
            return
        # Prefer the result with the highest smeta total and valid completed total.
        def rank(item: Dict[str, Any]) -> Tuple[int, float, float]:
            return (
                1 if item.get("completed_total") else 0,
                float(item.get("smeta_total") or 0),
                float(item.get("completed_total") or 0),
            )
        best = sorted(results, key=rank, reverse=True)[0]
        smeta_total = best.get("smeta_total")
        completed_total = best.get("completed_total")

        if best.get("project_name"):
            parsed.project_name = best["project_name"]
        if best.get("currency"):
            parsed.currency = best["currency"]
        if smeta_total:
            parsed.total_cost = float(smeta_total)
            parsed.planned_cost = float(smeta_total)
        if completed_total and smeta_total and float(smeta_total) > 0:
            completed = float(completed_total)
            upper = float(smeta_total) * 1.2
            if 0 < completed <= upper:
                parsed.actual_cost = round(completed, 2)
                parsed.actual_execution = round(self._clamp((completed / float(smeta_total)) * 100, 0, 100), 2)
                parsed.cost_variance_percent = None
                parsed.evidence["actual_execution_source"] = "Azerbaijani F-2 parser: completed amount / Nokopitelni smeta total"
            else:
                parsed.evidence["rejected_az_f2_completed_amount"] = completed
                parsed.warnings.append("Azerbaijani F-2 completed amount exceeded 120% of smeta total and requires manual mapping confirmation.")
        parsed.evidence["az_f2_parser"] = {
            "smeta_total": smeta_total,
            "completed_total": completed_total,
            "actual_execution": best.get("actual_execution"),
            "periods": best.get("periods") or [],
            "smeta_sheet": best.get("smeta_sheet"),
            "source": best.get("source"),
        }
        if best.get("periods"):
            parsed.evidence["f2_periods"] = best["periods"]
        self._clear_actual_cost_rejection_if_valid(parsed)
        if best.get("section_breakdown"):
            parsed.evidence["section_breakdown"] = best["section_breakdown"]

    def _try_az_f2_smeta_parse(self, path: Path) -> Optional[Dict[str, Any]]:
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
        except Exception:
            return None

        sheet_names = wb.sheetnames
        norm_names = [self._norm(s) for s in sheet_names]

        smeta_ws = None
        smeta_sheet_name = None
        for i, nm in enumerate(norm_names):
            if any(kw in nm for kw in self._AZ_SMETA_SHEET_NAMES):
                smeta_ws = wb[sheet_names[i]]
                smeta_sheet_name = sheet_names[i]
                break
        if smeta_ws is None:
            return None

        smeta_rows = [list(r) for r in smeta_ws.iter_rows(max_row=700, values_only=True)]
        smeta_rows = [r for r in smeta_rows if any(v not in (None, "") for v in r)]
        if not smeta_rows:
            return None

        project_name = self._extract_az_project_name_from_smeta(smeta_rows)
        currency = self._detect_az_currency_from_rows(smeta_rows[:30])
        smeta_total, smeta_total_row = self._find_az_smeta_total(smeta_rows)
        f2_cols, remaining_cols, header_row_idx = self._find_az_f2_columns(smeta_rows)

        result: Dict[str, Any] = {
            "project_name": project_name,
            "smeta_total": smeta_total,
            "currency": currency,
            "completed_total": None,
            "actual_execution": None,
            "periods": [],
            "section_breakdown": [],
            "smeta_sheet": smeta_sheet_name,
            "source": None,
        }

        if f2_cols and smeta_total and smeta_total > 0:
            completed, periods = self._sum_az_f2_columns(smeta_rows, f2_cols, header_row_idx, smeta_total)
            if completed and completed > 0:
                result["completed_total"] = completed
                result["actual_execution"] = round(min(100.0, (completed / smeta_total) * 100), 2)
                result["periods"] = periods
                result["source"] = "Nokopitelni F-2 period columns"

        if (result["completed_total"] is None or result["completed_total"] == 0) and smeta_total:
            f2_sheets = [sheet_names[i] for i, nm in enumerate(norm_names)
                         if self._AZ_F2_SHEET_PATTERN.search(sheet_names[i]) and sheet_names[i] != smeta_sheet_name]
            if f2_sheets:
                completed, periods = self._sum_az_separate_f2_sheets(wb, f2_sheets, smeta_total)
                if completed and completed > 0:
                    result["completed_total"] = completed
                    result["periods"] = periods
                    result["source"] = "Separate F-2 sheets"
                    result["actual_execution"] = round(min(100.0, (completed / smeta_total) * 100), 2)

        if smeta_total:
            result["section_breakdown"] = self._extract_az_section_breakdown(smeta_rows)
            return result
        return None

    def _extract_az_project_name_from_smeta(self, rows: List[List[Any]]) -> Optional[str]:
        # Prefer the 2nd visible row's first long text, which is where many local
        # Nokopitelni files keep the real object title.
        prioritized_rows = []
        if len(rows) > 1:
            prioritized_rows.append(rows[1])
        prioritized_rows.extend(rows[:8])
        for row in prioritized_rows:
            for val in row:
                if not val or not isinstance(val, str):
                    continue
                text = re.sub(r"\s+", " ", val).strip()
                if len(text) <= 20:
                    continue
                norm = self._norm(text)
                if any(bad in norm for bad in ("bolme 1", "torpaq isleri", "nasosxana", "yekun cem")):
                    continue
                construction_kws = (
                    "layihe", "tikinti", "abadliq", "insaat", "construction", "project",
                    "seheri", "susa", "merkezi", "kuce", "qiymet", "muqavile", "smeta",
                )
                if any(kw in norm for kw in construction_kws):
                    return text.replace("\n", " ").replace("\r", " ")[:180]
        return None

    def _detect_az_currency_from_rows(self, rows: List[List[Any]]) -> Optional[str]:
        text = " ".join(str(v) for row in rows for v in row if v).lower()
        if re.search(r"\b(azn|manat)\b|₼", text):
            return "AZN"
        if re.search(r"\b(usd|dollar)\b|\$", text):
            return "USD"
        if re.search(r"\b(eur|euro)\b|€", text):
            return "EUR"
        if any(ch in text for ch in "əğışöüç"):
            return "AZN"
        return None

    def _find_az_smeta_total(self, rows: List[List[Any]]) -> Tuple[Optional[float], Optional[int]]:
        pattern = re.compile("|".join(self._AZ_SMETA_TOTAL_PATTERNS), re.IGNORECASE)
        for ri, row in enumerate(rows):
            row_text_norm = self._norm(" ".join(str(v) for v in row if v))
            if pattern.search(row_text_norm):
                numbers = []
                for v in row:
                    if isinstance(v, (int, float)) and 10000 <= float(v) <= MAX_REASONABLE_MONEY:
                        numbers.append(float(v))
                    elif isinstance(v, str):
                        n = self._parse_number(v)
                        if n is not None and 10000 <= n <= MAX_REASONABLE_MONEY:
                            numbers.append(float(n))
                if numbers:
                    # In Nokopitelni total row, the baseline/smeta is usually the first
                    # large amount. Taking max may pick cumulative unrelated values.
                    return numbers[0], ri
        return None, None

    def _find_az_f2_columns(self, rows: List[List[Any]]) -> Tuple[List[int], List[int], Optional[int]]:
        f2_pattern = re.compile("|".join(self._AZ_F2_COL_PATTERNS), re.IGNORECASE)
        rem_pattern = re.compile("|".join(self._AZ_REMAINING_COL_PATTERNS), re.IGNORECASE)
        best: Tuple[List[int], List[int], Optional[int]] = ([], [], None)
        for ri, row in enumerate(rows[:30]):
            f2_cols, rem_cols = [], []
            for ci, val in enumerate(row):
                if not val:
                    continue
                norm_val = self._norm(str(val)).replace("-", " ")
                if f2_pattern.search(norm_val):
                    f2_cols.append(ci)
                elif rem_pattern.search(norm_val):
                    rem_cols.append(ci)
            if len(f2_cols) > len(best[0]):
                best = (f2_cols, rem_cols, ri)
        return best

    def _sum_az_f2_columns(
        self,
        rows: List[List[Any]],
        f2_cols: List[int],
        header_row_idx: Optional[int],
        smeta_total: float,
    ) -> Tuple[Optional[float], List[Dict[str, Any]]]:
        total_row_idx = self._find_az_smeta_total(rows)[1]
        if total_row_idx is None or not f2_cols:
            return None, []
        total_row = list(rows[total_row_idx])
        header = rows[header_row_idx] if header_row_idx is not None else []
        periods: List[Dict[str, Any]] = []
        completed_total = 0.0
        # Local Nokopitelni structure: each F-2 header is followed by a merged group;
        # the total row value for that period is typically at header column +2.
        offsets = (2, 1, 0, 3)
        used_data_cols = set()
        for f2_ci in f2_cols:
            chosen_val = None
            chosen_data_ci = None
            for offset in offsets:
                data_ci = f2_ci + offset
                if data_ci in used_data_cols or data_ci >= len(total_row):
                    continue
                val = self._parse_number(total_row[data_ci])
                if val is None:
                    continue
                if 100 <= val <= smeta_total * 1.2:
                    chosen_val = float(val)
                    chosen_data_ci = data_ci
                    break
            if chosen_val is None:
                continue
            used_data_cols.add(chosen_data_ci)
            col_name = str(header[f2_ci]).strip() if f2_ci < len(header) and header[f2_ci] else f"F-2 period {len(periods)+1}"
            periods.append({"period": col_name, "amount": round(chosen_val, 2), "data_column": self._excel_column_name(chosen_data_ci + 1)})
            completed_total += chosen_val
        if completed_total <= 0 or completed_total > smeta_total * 1.2:
            return None, periods
        return round(completed_total, 2), periods

    def _sum_az_separate_f2_sheets(self, wb: Any, f2_sheet_names: List[str], smeta_total: float) -> Tuple[Optional[float], List[Dict[str, Any]]]:
        total = 0.0
        periods: List[Dict[str, Any]] = []
        total_kws = re.compile(r"yekun|c[əe]mi|total|sum|cemi|yekun\s+c[əe]m|yerine yetirilmis|icra", re.IGNORECASE)
        for sname in f2_sheet_names:
            try:
                ws = wb[sname]
                sheet_rows = [list(r) for r in ws.iter_rows(max_row=300, values_only=True)]
            except Exception:
                continue
            sheet_total: Optional[float] = None
            source = None
            for row in reversed(sheet_rows):
                row_text = self._norm(" ".join(str(v) for v in row if v))
                if total_kws.search(row_text):
                    nums = []
                    for v in row:
                        n = self._parse_number(v)
                        if n is not None and 100 <= n <= smeta_total * 1.2:
                            nums.append(float(n))
                    if nums:
                        sheet_total = max(nums)
                        source = self._row_text(row)[:180]
                        break
            if sheet_total and 100 <= sheet_total <= smeta_total * 1.2:
                total += sheet_total
                periods.append({"period": sname, "amount": round(sheet_total, 2), "source": source})
        if total <= 0 or total > smeta_total * 1.2:
            return None, periods
        return round(total, 2), periods

    def _extract_az_section_breakdown(self, rows: List[List[Any]]) -> List[Dict[str, Any]]:
        section_pattern = re.compile(r"^(bolme|bölmə|section|part|chapter|hisse|hissə)\s*\d", re.IGNORECASE)
        sections: List[Dict[str, Any]] = []
        for row in rows:
            if not row:
                continue
            label = ""
            for cell in row[:4]:
                if isinstance(cell, str) and section_pattern.match(self._norm(cell)):
                    label = cell.strip()
                    break
            if not label:
                continue
            nums = []
            for v in row:
                n = self._parse_number(v)
                if n is not None and n > 100:
                    nums.append(float(n))
            if nums:
                sections.append({"section": label[:100], "amount": max(nums)})
        return sections[:30]


    def _clear_actual_cost_rejection_if_valid(self, parsed: ParsedProjectData) -> None:
        """Remove stale generic actual-cost rejection warnings when a deterministic
        F-2/Nokopitelni calculation has already produced a valid actual cost.

        This happens when generic column scanning first sees an impossible number,
        then the Azerbaijan F-2 parser correctly validates the completed amount.
        The final dashboard should not warn about rejected_actual_cost if the
        accepted actual_cost is within the 120% smeta safety threshold.
        """
        try:
            if not parsed.total_cost or parsed.actual_cost is None:
                return
            total = float(parsed.total_cost)
            actual = float(parsed.actual_cost)
            if total <= 0 or not (0 <= actual <= total * 1.2):
                return
        except Exception:
            return

        stale_keys = {
            "rejected_actual_cost",
            "rejected_actual_cost_postprocess",
            "rejected_f2_completed_amount",
            "rejected_az_f2_completed_amount",
        }
        for key in stale_keys:
            parsed.evidence.pop(key, None)

        stale_phrases = (
            "exceeded 120%",
            "exceeds 120%",
            "requires manual mapping confirmation",
            "Manual confirmation is required",
        )
        parsed.warnings = [
            w for w in parsed.warnings
            if not any(phrase.lower() in str(w).lower() for phrase in stale_phrases)
        ]

    def _apply_commercial_accuracy_guardrails(self, parsed: ParsedProjectData) -> None:
        """Protect commercial KPIs from unconfirmed over-baseline values.

        In construction payment/cost files, an extracted amount higher than the
        smeta/contract total is not automatically wrong, but it is not safe to
        use as a final KPI without user confirmation. It can indicate VAT,
        duplicate cumulative totals, approved variations, or a wrong mapped
        amount column.
        """
        try:
            if not parsed.total_cost or parsed.actual_cost is None:
                return
            total = float(parsed.total_cost)
            actual = float(parsed.actual_cost)
            if total <= 0 or actual <= total * 1.005:
                return
        except Exception:
            return

        parsed.evidence["commercial_guardrail"] = "actual_cost_exceeds_smeta_total"
        parsed.evidence["needs_confirmation_actual_cost"] = {
            "detected_actual_cost": round(actual, 2),
            "smeta_total": round(total, 2),
            "ratio_percent": round((actual / total) * 100, 2),
            "reason": "Detected actual completed cost exceeds the smeta/contract baseline.",
        }
        parsed.warnings.append(
            "Actual completed cost exceeds the detected smeta/contract total and needs mapping confirmation before commercial use."
        )
        parsed.actual_cost = None
        parsed.actual_execution = None
        parsed.cost_variance_percent = None


    def _extract_workforce_productivity(self, parsed: ParsedProjectData, rows: Sequence[Sequence[Any]]) -> None:
        """Attach workforce productivity planning evidence when activity rows are present."""
        try:
            result = analyze_workforce_productivity(rows)
        except Exception as exc:
            parsed.warnings.append(f"Workforce productivity analysis could not run: {exc}")
            return
        activities = result.get("activities") or []
        if not activities:
            return
        parsed.evidence["workforce_productivity"] = result
        summary = result.get("summary") or {}
        if summary.get("total_actual_workers") is not None and parsed.workforce_current is None:
            parsed.workforce_current = int(summary["total_actual_workers"])
        if summary.get("total_required_workers") is not None and parsed.workforce_required is None:
            parsed.workforce_required = int(summary["total_required_workers"])
        for warning in result.get("warnings") or []:
            if warning not in parsed.warnings:
                parsed.warnings.append(warning)

    def _post_process(self, parsed: ParsedProjectData) -> None:
        if parsed.delay_days is not None and parsed.delay_days < 0:
            parsed.evidence["negative_delay_days_normalized"] = parsed.delay_days
            parsed.delay_days = 0

        for field in ("workforce_current", "workforce_required"):
            value = getattr(parsed, field, None)
            if value is not None and int(value) < 0:
                parsed.evidence[f"negative_{field}_rejected"] = value
                setattr(parsed, field, None)

        if parsed.delay_days is None and parsed.baseline_finish and parsed.estimated_finish:
            b = self._parse_date(parsed.baseline_finish)
            e = self._parse_date(parsed.estimated_finish)
            if b and e:
                parsed.delay_days = (e - b).days

        if parsed.total_cost and parsed.actual_cost and float(parsed.actual_cost) > float(parsed.total_cost) * 1.2:
            parsed.evidence["rejected_actual_cost_postprocess"] = parsed.actual_cost
            parsed.warnings.append("Actual cost was not used because it exceeded 120% of the detected smeta total.")
            parsed.actual_cost = None
            parsed.cost_variance_percent = None

        # Commercial accuracy guardrail: if the detected actual completed amount
        # is higher than the detected smeta/contract baseline, do not present it
        # as a confirmed commercial KPI. It may be a VAT-inclusive value, duplicated
        # cumulative total, approved variation, or wrong mapped column, so it must
        # be confirmed in the mapping screen first.
        self._apply_commercial_accuracy_guardrails(parsed)

        if parsed.planned_cost and parsed.actual_cost is not None and (parsed.cost_variance_percent is None or parsed.evidence.get("authoritative_cost_pair")):
            parsed.cost_variance_percent = round(((float(parsed.actual_cost) - float(parsed.planned_cost)) / float(parsed.planned_cost)) * 100, 2)

        self._clear_actual_cost_rejection_if_valid(parsed)

        if parsed.planned_execution is not None:
            parsed.planned_execution = self._clamp(round(parsed.planned_execution, 2), 0, 100)
        if parsed.actual_execution is not None:
            parsed.actual_execution = self._clamp(round(parsed.actual_execution, 2), 0, 100)

        if not parsed.sheets:
            parsed.warnings.append("No readable sheets or supported file structures were found.")
        if parsed.planned_execution is None and parsed.actual_execution is None:
            parsed.warnings.append("Plan/fact progress could not be confidently extracted from the uploaded data.")
        if self.analysis_type in {"cost", "progress", "all"} and parsed.cost_variance_percent is None and parsed.planned_cost is None and parsed.actual_cost is None:
            parsed.warnings.append("Cost variance could not be calculated because cost fields were not clearly mapped.")
        if parsed.currency is None:
            parsed.currency = "AZN" if parsed.language_hint in {"az", "tr"} else None

    def _project_candidates_from_rows(self, rows: Sequence[Sequence[Any]]) -> List[str]:
        candidates: List[str] = []
        for row_pos, row in enumerate(rows):
            values = [str(v).strip() for v in row if v not in (None, "")]
            if not values:
                continue
            row_text = " | ".join(values)
            norm = self._norm(row_text)

            # Many Azerbaijani smeta/nakopitelni workbooks keep the real project title
            # as a long merged cell on the 2nd visible row. Capture it before section rows
            # such as "Bölmə 1. Torpaq işləri ..." can pollute the candidate list.
            if row_pos == 1:
                first_value = values[0]
                if self._looks_like_project_name(first_value) and len(first_value) >= 18:
                    candidates.append(first_value)

            for idx, val in enumerate(values):
                nval = self._norm(val)
                if any(label in nval for label in PROJECT_LABELS):
                    if idx + 1 < len(values) and self._looks_like_project_name(values[idx + 1]):
                        candidates.append(values[idx + 1])
                    elif ":" in val:
                        after = val.split(":", 1)[1].strip()
                        if self._looks_like_project_name(after):
                            candidates.append(after)

            if self._looks_like_project_name(row_text) and any(k in norm for k in ("susa", "merkezi", "mərkəzi", "abadliq", "abadlıq", "layihe", "layihə", "project", "construction")):
                candidates.append(row_text)
        return candidates

    def _choose_project_name(self, candidates: Sequence[str], paths: Sequence[Path]) -> str:
        cleaned: List[str] = []
        for c in candidates:
            value = re.sub(r"\s+", " ", str(c or "")).strip(" -–—:|")
            if self._looks_like_project_name(value):
                cleaned.append(value[:180])
        if cleaned:
            unique = list(dict.fromkeys(cleaned))
            unique.sort(key=lambda x: (self._project_name_score(x), min(len(x), 140), -x.count("|")), reverse=True)
            return unique[0]
        for path in paths:
            stem = re.sub(r"[_-]+", " ", path.stem).strip()
            norm_stem = self._norm(stem)
            if any(token in norm_stem for token in ("template", "devbareun professional upload", "cost template", "schedule template", "workforce template")):
                continue
            if self._looks_like_project_name(stem):
                return stem[:120]
        return "DevBareun Uploaded Project"

    def _project_name_score(self, text: str) -> int:
        norm = self._norm(text)
        score = 0
        # Strongly prefer the real object/title line used in Azerbaijani construction
        # estimates over section titles such as "Bölmə 1".
        weighted_tokens = {
            "susa": 8,
            "seheri": 4,
            "merkezi": 7,
            "kuce": 5,
            "abadliq": 8,
            "isleri": 2,
            "layihe": 3,
            "project": 3,
            "tikinti": 2,
            "construction": 2,
            "massiv": 2,
            "bina": 1,
            "residential": 2,
        }
        for token, weight in weighted_tokens.items():
            if token in norm:
                score += weight
        for bad_token, penalty in {
            "bolme": 25,
            "torpaq isleri": 10,
            "nasosxana": 8,
            "yekun": 8,
            "cedvel": 6,
            "f 2": 10,
            "forma 2": 10,
        }.items():
            if bad_token in norm:
                score -= penalty
        if "|" in text:
            score -= 6
        if any(label in norm for label in ("layihenin adi", "project name")):
            score -= 2
        return score

    def _looks_like_project_name(self, value: str) -> bool:
        value = str(value or "").strip()
        if len(value) < 4 or len(value) > 240:
            return False
        norm = self._norm(value)
        bad = {"sheet1", "upload", "data", "cost", "schedule", "progress", "report", "dashboard"}
        template_sheet_phrases = {
            "devbareun professional construction upload template",
            "full dashboard input",
            "cost estimate smeta",
            "f2 progress payment",
            "schedule plan actual",
            "workforce productivity",
            "equipment usage",
            "lists",
            "readme",
            "use this workbook",
            "supported analysis",
            "recommended workflow",
            "core rule",
            "important",
        }
        if norm in bad or any(phrase in norm for phrase in template_sheet_phrases):
            return False
        if re.fullmatch(r"[\d\s.,:/-]+", value):
            return False
        return True

    def _language_hint(self, text: str) -> Optional[str]:
        lower = text.lower()
        if any(ch in text for ch in AZ_TR_CHARS) or any(k in lower for k in ("layih", "icra", "faktiki", "məbləğ", "smeta")):
            return "az"
        if any(k in lower for k in ("maliyet", "hakediş", "iş programı")):
            return "tr"
        return "en" if re.search(r"\b(project|schedule|cost|actual|planned)\b", lower) else None

    def _detect_currency(self, text: str, language_hint: Optional[str]) -> Optional[str]:
        lower = text.lower()
        if re.search(r"\b(azn|manat)\b|₼", lower):
            return "AZN"
        if re.search(r"\b(usd|dollar)\b|\$", lower):
            return "USD"
        if re.search(r"\b(eur|euro)\b|€", lower):
            return "EUR"
        if re.search(r"\btry\b|₺|tl\b", lower):
            return "TRY"
        if language_hint in {"az", "tr"}:
            return "AZN"
        return None

    def _extract_percent_by_labels(self, text: str, labels: Sequence[str]) -> Optional[float]:
        norm_text = self._norm(text)
        # Work on original lines to avoid losing % signs.
        for line in text.splitlines():
            nline = self._norm(line)
            if any(self._norm(label) in nline for label in labels):
                value = self._parse_percent(line)
                if value is not None:
                    return value
        # Regex fallback for same-line label: number + percent within 50 chars.
        for label in labels:
            pattern = re.compile(re.escape(label), re.IGNORECASE)
            match = pattern.search(text)
            if match:
                window = text[match.start() : match.start() + 120]
                value = self._parse_percent(window)
                if value is not None:
                    return value
        return None

    def _extract_int_by_labels(self, text: str, labels: Sequence[str]) -> Optional[int]:
        for line in text.splitlines():
            nline = self._norm(line)
            if any(self._norm(label) in nline for label in labels):
                value = self._first_int([line])
                if value is not None:
                    return value
        return None

    def _extract_money_by_labels(self, text: str, labels: Sequence[str]) -> Optional[float]:
        for line in text.splitlines():
            nline = self._norm(line)
            if any(self._norm(label) in nline for label in labels):
                value = self._parse_number(line)
                if value is not None:
                    return value
        return None

    def _extract_date_by_labels(self, text: str, labels: Sequence[str]) -> Optional[str]:
        for line in text.splitlines():
            nline = self._norm(line)
            if any(self._norm(label) in nline for label in labels):
                value = self._first_date([line])
                if value:
                    return value
        return None

    def _first_percent(self, values: Sequence[Any]) -> Optional[float]:
        for value in values:
            result = self._parse_percent(value)
            if result is not None:
                return result
        return None

    def _first_int(self, values: Sequence[Any]) -> Optional[int]:
        for value in values:
            number = self._parse_number(value)
            if number is not None and abs(number) < 100000:
                return int(round(number))
        return None

    def _first_date(self, values: Sequence[Any]) -> Optional[str]:
        for value in values:
            parsed = self._parse_date(value)
            if parsed:
                return parsed.isoformat()
        return None

    def _parse_percent(self, value: Any) -> Optional[float]:
        if value is None or value == "":
            return None
        if isinstance(value, (int, float)):
            number = float(value)
            if 0 <= number <= 1:
                return round(number * 100, 2)
            if 0 <= number <= 100:
                return round(number, 2)
            return None
        text = str(value)
        matches = re.findall(r"[-+]?\d+(?:[.,]\d+)?\s*%?", text)
        for match in matches:
            has_percent = "%" in match or "%" in text
            number = self._parse_number(match)
            if number is None:
                continue
            if has_percent or 0 <= number <= 100:
                if 0 <= number <= 1 and not has_percent:
                    number *= 100
                if 0 <= number <= 100:
                    return round(number, 2)
        return None

    def _parse_number(self, value: Any) -> Optional[float]:
        if value is None or value == "":
            return None
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip()
        if not text:
            return None
        # Remove currency, spaces and words, keep numeric punctuation.
        cleaned = re.sub(r"[^0-9,\.\-+]", "", text)
        if not cleaned or cleaned in {"-", "+", ".", ","}:
            return None
        # Handle both 1,234.56 and 1.234,56 formats.
        if "," in cleaned and "." in cleaned:
            if cleaned.rfind(",") > cleaned.rfind("."):
                cleaned = cleaned.replace(".", "").replace(",", ".")
            else:
                cleaned = cleaned.replace(",", "")
        elif "," in cleaned:
            parts = cleaned.split(",")
            if len(parts[-1]) in {1, 2}:
                cleaned = cleaned.replace(",", ".")
            else:
                cleaned = cleaned.replace(",", "")
        try:
            number = float(cleaned)
            # Guardrail for MVP: construction line totals above this level are usually
            # parsing errors caused by concatenated text/numeric cells. Enterprise-scale
            # projects can still be handled through manual mapping confirmation.
            if abs(number) > MAX_REASONABLE_MONEY:
                return None
            return number
        except ValueError:
            return None

    def _parse_date(self, value: Any) -> Optional[date]:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if value is None or value == "":
            return None
        text = str(value)
        # Avoid parsing plain amounts/percentages as dates.
        if not re.search(r"\d{1,4}[./-]\d{1,2}[./-]\d{1,4}", text):
            return None
        try:
            return date_parser.parse(text, dayfirst=True, fuzzy=True).date()
        except Exception:
            return None

    def _clean_cell(self, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, str):
            return re.sub(r"\s+", " ", value).strip()
        return value

    def _row_text(self, row: Sequence[Any]) -> str:
        return " | ".join(str(v) for v in row if v not in (None, ""))

    def _norm(self, value: str) -> str:
        text = str(value or "").lower()
        replacements = {
            "ə": "e", "ı": "i", "ğ": "g", "ü": "u", "ö": "o", "ş": "s", "ç": "c",
            "Ə": "e", "İ": "i", "Ğ": "g", "Ü": "u", "Ö": "o", "Ş": "s", "Ç": "c",
        }
        for src, dst in replacements.items():
            text = text.replace(src, dst)
        text = re.sub(r"[^a-z0-9%+./\- ]+", " ", text)
        return re.sub(r"\s+", " ", text).strip()

    def _has_any(self, text: str, labels: Sequence[str]) -> bool:
        return any(self._norm(label) in text for label in labels)

    def _excel_column_name(self, index: int) -> str:
        name = ""
        while index:
            index, rem = divmod(index - 1, 26)
            name = chr(65 + rem) + name
        return name

    def _clamp(self, value: float, low: float, high: float) -> float:
        return max(low, min(high, value))
